import asyncio
import base64
import os
import tempfile
import time
import unittest
import zipfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

import excel_image_server as server


class DummyUpload:
    def __init__(self, filename, data):
        self.filename = filename
        self.file = BytesIO(data)


class DummyForm:
    def __init__(self, fields):
        self.fields = fields

    def get(self, key, default=None):
        return self.fields.get(key, default)

    def getlist(self, key):
        value = self.fields.get(key, [])
        return value if isinstance(value, list) else [value]

    def close(self):
        return None


class DummyRequest:
    def __init__(self, fields, progress_id):
        self._form = DummyForm(fields)
        self.query_params = {"progress_id": progress_id}

    async def form(self, **_kwargs):
        return self._form


async def collect_streaming_response(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


class UrlMappingTests(unittest.TestCase):
    def setUp(self):
        self.old_tempdir = tempfile.tempdir
        self.old_temp_env = {
            name: os.environ.get(name)
            for name in ("TEMP", "TMP", "TMPDIR")
        }
        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        server.configure_runtime(
            {
                "host": "127.0.0.1",
                "port": 8091,
                "xml_url": "http://example.test/images.xml",
                "images_dir": root / "Foto",
                "images_base_url": "http://images.example.test/foto/",
                "work_dir": root / "work",
                "public_log_file": root / "log.txt",
                "max_log_lines": 100,
                "max_upload_mb": 10,
                "hard_max_output_mb": 100,
                "default_desired_output_mb": 20,
                "max_parallel_jobs": 1,
                "job_ttl_minutes": 120,
                "xml_cache_seconds": 300,
                "default_start_row": 2,
                "default_article_column": "A",
                "default_image_column": "B",
                "target_image_width_px": 120,
                "adjust_row_height": True,
                "default_cell_background_color": "D9D9D9",
                "image_padding_px": 1,
                "image_width_guard_px": 6,
                "jpeg_quality": 92,
                "min_jpeg_quality": 82,
                "download_timeout_seconds": 5,
            }
        )
        (root / "Foto" / "BABY TEAM").mkdir(parents=True)

    def tearDown(self):
        tempfile.tempdir = self.old_tempdir
        for name, value in self.old_temp_env.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value
        self.temp_dir.cleanup()

    def test_public_url_maps_to_local_path(self):
        result = server.image_url_to_local_path(
            "http://images.example.test/foto/BABY%20TEAM/0010.jpg"
        )
        self.assertTrue(str(result).endswith(r"Foto\BABY TEAM\0010.jpg") or str(result).endswith("Foto/BABY TEAM/0010.jpg"))

    def test_double_slash_base_url_is_normalized_for_local_image_mapping(self):
        server.CONFIG["images_base_url"] = server.normalize_base_url(
            "http://images.example.test//foto"
        )

        result = server.image_url_to_local_path(
            "http://images.example.test/foto/BABY%20TEAM/0010.jpg"
        )

        self.assertTrue(str(result).endswith(r"Foto\BABY TEAM\0010.jpg") or str(result).endswith("Foto/BABY TEAM/0010.jpg"))

    def test_other_base_url_is_not_mapped(self):
        self.assertIsNone(server.image_url_to_local_path("http://example.test/0010.jpg"))

    def test_excel_base64_is_read_into_memory_with_neutral_metadata(self):
        encoded = base64.b64encode(b"not a real workbook").decode("ascii")

        buffer, suffix, size = server.read_excel_base64_to_memory(
            encoded,
            "secret payroll clients.xlsx",
        )

        self.assertEqual(suffix, ".xlsx")
        self.assertEqual(size, len(b"not a real workbook"))
        self.assertEqual(buffer.getvalue(), b"not a real workbook")
        self.assertFalse((Path(self.temp_dir.name) / "secret payroll clients.xlsx").exists())
        buffer.close()

    def test_python_temp_dir_is_inside_work_dir(self):
        expected = server.CONFIG["work_dir"] / "python_temp"

        self.assertEqual(Path(tempfile.tempdir), expected)
        self.assertEqual(Path(os.environ["TEMP"]), expected)
        self.assertEqual(Path(os.environ["TMP"]), expected)
        self.assertTrue(expected.is_dir())

    def test_missing_python_temp_dir_is_recreated_before_job(self):
        expected = server.CONFIG["work_dir"] / "python_temp"
        expected.rmdir()

        job_dir = server.make_job_dir("missing_temp")

        self.assertTrue(expected.is_dir())
        self.assertEqual(Path(tempfile.tempdir), expected)
        self.assertTrue(job_dir.is_dir())

    def test_configure_runtime_removes_stale_job_dirs(self):
        stale_job = server.CONFIG["work_dir"] / "jobs" / "process_stale"
        stale_job.mkdir(parents=True)
        (stale_job / "input.xlsx").write_bytes(b"secret")

        server.configure_runtime(dict(server.CONFIG))

        self.assertFalse(stale_job.exists())

    def test_cleanup_old_jobs_removes_empty_process_dirs_immediately(self):
        empty_job = server.CONFIG["work_dir"] / "jobs" / "process_empty"
        empty_job.mkdir(parents=True)

        server.cleanup_old_jobs()

        self.assertFalse(empty_job.exists())

    def test_work_dir_must_not_be_inside_public_sharefiles_tree(self):
        with self.assertRaises(ValueError):
            server.validate_private_work_dir(
                Path(r"D:\ExamplePublicRoot\private_work"),
                [
                    Path(r"D:\ExamplePublicRoot\public\xml"),
                    Path(r"D:\ExamplePublicRoot\public\foto"),
                ],
            )

    def test_log_trim_does_not_create_public_tmp_file(self):
        server.MAX_LOG_LINES = 2
        server.LOG_FILE.write_text("one\ntwo\nthree\n", encoding="utf-8-sig")

        server.trim_log_file()

        self.assertEqual(
            server.LOG_FILE.read_text(encoding="utf-8-sig").splitlines(),
            ["two", "three"],
        )
        self.assertEqual(list(server.LOG_FILE.parent.glob("*.tmp")), [])

    def test_prepare_public_log_preserves_recent_lines(self):
        server.MAX_LOG_LINES = 2
        server.LOG_FILE.write_text("one\ntwo\nthree\n", encoding="utf-8-sig")

        server.prepare_public_log()

        self.assertEqual(
            server.LOG_FILE.read_text(encoding="utf-8-sig").splitlines(),
            ["two", "three"],
        )

    def test_configure_runtime_does_not_clear_public_log(self):
        server.LOG_FILE.write_text("old line\n", encoding="utf-8-sig")

        server.configure_runtime(dict(server.CONFIG))

        self.assertEqual(
            server.LOG_FILE.read_text(encoding="utf-8-sig").splitlines(),
            ["old line"],
        )

    def test_result_filename_keeps_original_name(self):
        self.assertEqual(
            server.make_download_filename(
                r"C:\fakepath\Отчет клиентов.xlsx",
                ".xlsx",
            ),
            "Отчет клиентов_with Images.xlsx",
        )

    def test_log_is_exposed_only_through_caddy(self):
        route_paths = {route.path for route in server.app.routes}
        self.assertNotIn("/log", route_paths)
        page = server.render_page()
        self.assertNotIn("excel_image_server.log", page)
        self.assertNotIn("Стан XML", page)

    def test_attached_file_can_be_cleared_without_submitting_form(self):
        page = server.render_page()

        self.assertIn('id="clearFileButton"', page)
        self.assertIn('type="button"', page)
        self.assertIn("resetAttachedFile();", server.CLIENT_JS)

    def test_upload_hint_explains_external_excel_links(self):
        page = server.render_page()

        self.assertIn("ВПР", page)
        self.assertIn("XLOOKUP", page)
        self.assertIn("через відновлення", page)
        self.assertIn("зовнішні посилання можуть втратитися", page)
        self.assertIn("збережені значення мають лишитися", page)

    def test_workbooks_are_opened_with_external_links_preserved(self):
        with patch.object(server, "load_workbook") as loader:
            workbook = loader.return_value
            workbook.sheetnames = ["Sheet1"]

            result = server.get_sheet_names(Path("book.xlsx"))

        self.assertEqual(result, ["Sheet1"])
        self.assertTrue(loader.call_args.kwargs["keep_links"])

    def test_processing_report_ui_is_available(self):
        page = server.render_page()

        self.assertIn('id="processingReport"', page)
        self.assertIn('id="reportContent"', page)
        self.assertIn("renderProcessingReport", server.CLIENT_JS)

    def test_processing_can_be_cancelled_from_ui(self):
        page = server.render_page()

        self.assertIn('id="cancelButton"', page)
        self.assertIn('id="cancelButton" class="cancel-button" type="button" disabled', page)
        self.assertIn("/cancel/", server.CLIENT_JS)
        self.assertIn("/heartbeat/", server.CLIENT_JS)
        self.assertIn("cancelRequested", server.CLIENT_JS)
        self.assertIn("beforeunload", server.CLIENT_JS)
        self.assertIn("sendBeacon", server.CLIENT_JS)

    def test_cell_background_can_be_disabled_in_ui(self):
        page = server.render_page()

        self.assertIn('name="skip_cell_background"', page)
        self.assertIn("Не змінювати фон комірки", page)
        self.assertIn("skip_cell_background", server.CLIENT_JS)

    def test_desired_output_size_is_integer_from_1_to_100_in_ui(self):
        page = server.render_page()

        self.assertIn('name="desired_output_mb" type="number" min="1" max="100" step="1"', page)
        self.assertIn('value="20"', page)
        self.assertNotIn('desiredSizeHint', page)
        self.assertNotIn('applyRecommendedSizeButton', page)
        self.assertNotIn("recommendedSizeForImageRows", server.CLIENT_JS)

    def test_local_folder_picker_uses_styled_ukrainian_button(self):
        page = server.render_page()

        self.assertIn('class="folder-input"', page)
        self.assertIn('class="folder-picker-button"', page)
        self.assertIn('class="folder-picker-title"', page)
        self.assertIn('class="folder-picker-name"', page)
        self.assertIn('class="folder-clear-button"', page)
        self.assertIn("Вибрати папку", page)
        self.assertIn("перетягнути сюди", page)
        self.assertIn("Папку прикріплено", server.CLIENT_JS)
        self.assertIn("clearLocalFolder", server.CLIENT_JS)
        self.assertIn("is-attached", server.CLIENT_JS)
        self.assertIn("webkitGetAsEntry", server.CLIENT_JS)
        self.assertIn("filesFromDrop", server.CLIENT_JS)
        self.assertIn("getLocalFiles", server.CLIENT_JS)

    def test_sources_can_be_disabled_in_priority_selects(self):
        page = server.render_page()

        self.assertIn('<option value="off">Вимкнено</option>', page)
        self.assertIn("select.value !== 'off'", server.CLIENT_JS)
        self.assertEqual(
            server.validate_source_order('["local_1","local_2"]'),
            ["local_1", "local_2"],
        )

    def test_progress_can_include_processing_report(self):
        server.set_progress(
            "report-1",
            100,
            "done",
            status="done",
            inserted=2,
            total=2,
            report={
                "inserted": 2,
                "not_found_count": 1,
                "not_found": ["0010"],
                "failed_count": 0,
                "failed": [],
            },
        )

        snapshot = server.get_progress_snapshot("report-1")

        self.assertEqual(snapshot["report"]["inserted"], 2)
        self.assertEqual(snapshot["report"]["not_found"], ["0010"])

    def test_progress_cancel_request_is_recorded(self):
        server.set_progress("cancel-1", 20, "running")

        self.assertTrue(server.request_cancel("cancel-1"))
        snapshot = server.get_progress_snapshot("cancel-1")

        self.assertEqual(snapshot["status"], "cancelling")
        self.assertTrue(snapshot["cancel_requested"])

    def test_missing_heartbeat_requests_cancellation(self):
        server.set_progress("stale-1", 10, "running")
        server.mark_progress_requires_heartbeat("stale-1")
        with server.PROGRESS_LOCK:
            server.JOBS["stale-1"]["last_client_seen_monotonic"] = (
                time.monotonic() - server.CLIENT_HEARTBEAT_TIMEOUT_SECONDS - 1
            )

        self.assertTrue(server.is_cancel_requested("stale-1"))
        snapshot = server.get_progress_snapshot("stale-1")
        self.assertEqual(snapshot["status"], "cancelling")

    def test_heartbeat_survives_progress_updates(self):
        server.set_progress("heartbeat-1", 10, "running")
        server.mark_progress_requires_heartbeat("heartbeat-1")
        self.assertTrue(server.mark_client_heartbeat("heartbeat-1"))

        server.set_progress("heartbeat-1", 20, "still running")
        snapshot = server.get_progress_snapshot("heartbeat-1")

        self.assertTrue(snapshot["heartbeat_required"])
        self.assertIn("last_client_seen_at", snapshot)

    def test_requiring_heartbeat_marks_client_seen_now(self):
        server.set_progress("heartbeat-now", 0, "starting")
        server.mark_progress_requires_heartbeat("heartbeat-now")

        snapshot = server.get_progress_snapshot("heartbeat-now")

        self.assertTrue(snapshot["heartbeat_required"])
        self.assertIn("last_client_seen_at", snapshot)
        self.assertFalse(server.is_cancel_requested("heartbeat-now"))

    def test_original_local_article_matching_rules_are_preserved(self):
        self.assertEqual(
            server.normalize_local_article("AB.12", for_excel=True),
            "ab-12",
        )
        self.assertEqual(
            server.normalize_local_article("AB-12", for_excel=False),
            "ab-12",
        )
        self.assertEqual(
            server.normalize_local_article("10.0", for_excel=True),
            "10-0",
        )
        self.assertNotEqual(
            server.normalize_local_article("0010", for_excel=True),
            server.normalize_local_article("0010_1", for_excel=False),
        )
        self.assertNotEqual(
            server.normalize_local_article("1001", for_excel=True),
            server.normalize_local_article("1001_1", for_excel=False),
        )

    def test_xml_article_matching_treats_underscore_as_dash(self):
        self.assertEqual(
            server.normalize_article_key("F9578-G0175_F9582"),
            server.normalize_article_key("F9578-G0175-F9582"),
        )

    def test_local_upload_name_must_match_normalized_article(self):
        image = BytesIO()
        PILImage.new("RGB", (40, 40), color=(0, 0, 255)).save(
            image,
            format="JPEG",
        )
        job_dir = Path(self.temp_dir.name) / "local_job"
        job_dir.mkdir()
        upload = DummyUpload("AB-12.jpg", image.getvalue())
        manifest = '[{"source":"local_1","key":"ab-12"}]'

        result = server.save_local_image_uploads(
            [upload],
            manifest,
            job_dir,
        )

        self.assertIn("ab-12", result["local_1"])
        self.assertTrue(result["local_1"]["ab-12"].is_file())

    def test_reload_message_uses_valid_utf8(self):
        server.INDEX_META = {
            "loaded_at": "now",
            "products_count": 1,
            "images_count": 1,
        }
        with patch.object(server, "load_xml_index", return_value={}):
            page = server.reload_index()

        self.assertIn("Обробка Excel", page)
        self.assertNotIn("РёРЅРґРµРєСЃ", page)


class ExcelProcessingTests(unittest.TestCase):
    def setUp(self):
        self.old_tempdir = tempfile.tempdir
        self.old_temp_env = {
            name: os.environ.get(name)
            for name in ("TEMP", "TMP", "TMPDIR")
        }
        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        image_root = root / "Foto"
        image_root.mkdir()
        self.source_image_path = image_root / "0010.jpg"
        PILImage.new("RGB", (200, 100), color=(255, 0, 0)).save(
            self.source_image_path
        )

        server.configure_runtime(
            {
                "host": "127.0.0.1",
                "port": 8091,
                "xml_url": "http://example.test/images.xml",
                "images_dir": image_root,
                "images_base_url": "http://img.test/foto/",
                "work_dir": root / "work",
                "public_log_file": root / "log.txt",
                "max_log_lines": 100,
                "max_upload_mb": 10,
                "hard_max_output_mb": 100,
                "default_desired_output_mb": 20,
                "max_parallel_jobs": 1,
                "job_ttl_minutes": 120,
                "xml_cache_seconds": 300,
                "default_start_row": 2,
                "default_article_column": "A",
                "default_image_column": "B",
                "target_image_width_px": 120,
                "adjust_row_height": True,
                "default_cell_background_color": "D9D9D9",
                "image_padding_px": 1,
                "image_width_guard_px": 6,
                "jpeg_quality": 92,
                "min_jpeg_quality": 82,
                "download_timeout_seconds": 5,
            }
        )
        server.INDEX_CACHE = {"0010": ["http://img.test/foto/0010.jpg"]}
        server.INDEX_LOADED_AT = 10**9

    def tearDown(self):
        tempfile.tempdir = self.old_tempdir
        for name, value in self.old_temp_env.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value
        self.temp_dir.cleanup()

    def test_process_xlsx_inserts_one_image_and_keeps_formula(self):
        root = Path(self.temp_dir.name)
        source = root / "input.xlsx"
        output = root / "output.xlsx"

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        ws["A1"] = "article"
        ws["B1"] = "image"
        ws["C1"] = "formula"
        ws["A2"] = "0010"
        ws["C2"] = "=1+1"
        ws.column_dimensions["B"].width = 20
        workbook.save(source)

        info = server.inspect_workbook(source, preview_rows=2)
        self.assertEqual(info["sheets"], ["Sheet1"])
        self.assertEqual(info["previews"]["Sheet1"]["rows"][1][0], "0010")

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=5,
            cell_background_color="D9D9D9",
        )

        self.assertEqual(result["inserted"], 1)
        with zipfile.ZipFile(output) as archive:
            media_names = [
                name for name in archive.namelist() if name.startswith("xl/media/")
            ]
            self.assertEqual(len(media_names), 1)
            embedded_bytes = archive.read(media_names[0])
            self.assertEqual(embedded_bytes, self.source_image_path.read_bytes())
            with PILImage.open(BytesIO(embedded_bytes)) as embedded:
                self.assertEqual(embedded.size, (200, 100))

        saved = load_workbook(output)
        try:
            self.assertEqual(saved["Sheet1"]["C2"].value, "=1+1")
            self.assertEqual(len(saved["Sheet1"]._images), 1)
            self.assertGreater(saved["Sheet1"].row_dimensions[2].height, 0)
            self.assertTrue(saved["Sheet1"]["B2"].fill.fgColor.rgb.endswith("D9D9D9"))
            self.assertEqual(saved["Sheet1"]._images[0].anchor._from.colOff, pixels_to_EMU(1))
            self.assertEqual(saved["Sheet1"]._images[0].anchor._from.rowOff, pixels_to_EMU(1))
            expected_width = (
                server.column_width_to_pixels(saved["Sheet1"], "B")
                - server.CONFIG["image_padding_px"]
                - server.CONFIG["image_width_guard_px"]
            )
            self.assertEqual(
                saved["Sheet1"]._images[0].anchor.ext.cx,
                pixels_to_EMU(expected_width),
            )
            self.assertLess(
                saved["Sheet1"]._images[0].anchor.ext.cy,
                pixels_to_EMU(expected_width),
            )
        finally:
            saved.close()

    def test_xml_article_with_dash_matches_excel_article_with_underscore(self):
        root = Path(self.temp_dir.name)
        source = root / "xml_dash_input.xlsx"
        output = root / "xml_dash_output.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A2"] = "F9578-G0175_F9582"
        worksheet.column_dimensions["B"].width = 20
        workbook.save(source)
        workbook.close()

        server.INDEX_CACHE = {
            server.normalize_article_key("F9578-G0175-F9582"): [
                "http://img.test/foto/0010.jpg"
            ]
        }
        server.INDEX_LOADED_AT = 10**9

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=5,
            cell_background_color="D9D9D9",
        )

        self.assertEqual(result["inserted"], 1)
        self.assertEqual(result["not_found"], [])

    def test_process_xlsx_can_skip_cell_background_fill(self):
        root = Path(self.temp_dir.name)
        source = root / "no_fill_input.xlsx"
        output = root / "no_fill_output.xlsx"

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        ws["A2"] = "0010"
        ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        ws.column_dimensions["B"].width = 20
        workbook.save(source)

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=5,
            cell_background_color="D9D9D9",
            use_cell_background=False,
        )

        self.assertEqual(result["inserted"], 1)
        self.assertFalse(result["use_cell_background"])
        saved = load_workbook(output)
        try:
            self.assertEqual(saved["Sheet1"]["B2"].fill.fill_type, "solid")
            self.assertTrue(saved["Sheet1"]["B2"].fill.fgColor.rgb.endswith("FFFF00"))
        finally:
            saved.close()

    def test_desired_output_size_is_clamped_to_integer_range(self):
        root = Path(self.temp_dir.name)
        source = root / "size_clamp_input.xlsx"
        output = root / "size_clamp_output.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A2"] = "0010"
        workbook.save(source)
        workbook.close()

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=0.4,
            cell_background_color="D9D9D9",
        )

        self.assertEqual(result["desired_output_mb"], 1)

    def test_compression_preserves_full_image_resolution(self):
        source = PILImage.effect_noise((800, 600), 100).convert("RGB")
        encoded, width, height = server.encode_image_for_budget(
            source,
            target_bytes=50 * 1024,
        )

        self.assertLessEqual(len(encoded), 50 * 1024)
        self.assertEqual((width, height), (800, 600))
        with PILImage.open(BytesIO(encoded)) as compressed:
            self.assertEqual(compressed.size, (800, 600))

    def test_compression_downscales_when_quality_is_not_enough(self):
        source = PILImage.effect_noise((800, 600), 100).convert("RGB")
        encoded, width, height = server.encode_image_for_budget(
            source,
            target_bytes=3 * 1024,
        )

        self.assertLessEqual(len(encoded), 3 * 1024)
        self.assertLess(width, 800)
        self.assertLess(height, 600)

    def test_many_images_can_use_sub_8kb_budget(self):
        root = Path(self.temp_dir.name)
        source = root / "many_input.xlsx"
        output = root / "many_output.xlsx"
        image_path = root / "many_source.jpg"

        PILImage.effect_noise((320, 240), 100).convert("RGB").save(
            image_path,
            quality=95,
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.column_dimensions["B"].width = 20
        local_index = {}
        for row in range(2, 202):
            article = f"ART{row:04d}"
            worksheet.cell(row=row, column=1).value = article
            local_index[server.normalize_local_article(article, for_excel=True)] = image_path
        workbook.save(source)
        workbook.close()

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=1,
            cell_background_color="D9D9D9",
            local_images_by_source={
                "local_1": local_index,
                "local_2": {},
            },
            source_order=["local_1", "server", "local_2"],
        )

        self.assertEqual(result["inserted"], 200)
        self.assertLess(result["target_bytes_per_image"], 8 * 1024)
        self.assertLessEqual(result["output_size_mb"], 1.1)

    def test_process_excel_stops_when_cancelled(self):
        root = Path(self.temp_dir.name)
        source = root / "cancel_input.xlsx"
        output = root / "cancel_output.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A2"] = "0010"
        workbook.save(source)
        workbook.close()

        with self.assertRaises(server.ProcessingCancelled):
            server.process_excel(
                input_path=source,
                output_path=output,
                sheet_name="Sheet1",
                article_column="A",
                image_column="B",
                start_row=2,
                desired_output_mb=5,
                cell_background_color="D9D9D9",
                cancel_callback=lambda: (_ for _ in ()).throw(
                    server.ProcessingCancelled("cancelled")
                ),
            )

        self.assertFalse(output.exists())

    def test_insert_progress_matches_current_image_count(self):
        root = Path(self.temp_dir.name)
        source = root / "progress_input.xlsx"
        output = root / "progress_output.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.column_dimensions["B"].width = 20
        for row in range(2, 6):
            worksheet.cell(row=row, column=1).value = "0010"
        workbook.save(source)
        workbook.close()

        progress_events = []
        server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet1",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=5,
            cell_background_color="D9D9D9",
            progress_callback=lambda percent, message, **extra: progress_events.append(
                (percent, message, extra.get("inserted"), extra.get("total"))
            ),
        )

        insert_events = [
            event for event in progress_events
            if event[1] == "Вставка зображень у Excel..."
        ]
        self.assertIn((25.0, "Вставка зображень у Excel...", 1, 4), insert_events)
        self.assertIn((100.0, "Вставка зображень у Excel...", 4, 4), insert_events)

    def test_local_folder_can_override_server_by_priority(self):
        root = Path(self.temp_dir.name)
        source = root / "priority_input.xlsx"
        output = root / "priority_output.xlsx"
        local_image = root / "0010_local_source.jpg"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "article"
        worksheet["B1"] = "image"
        worksheet["A2"] = "0010"
        workbook.save(source)
        workbook.close()
        PILImage.new("RGB", (200, 100), color=(0, 0, 255)).save(local_image)

        result = server.process_excel(
            input_path=source,
            output_path=output,
            sheet_name="Sheet",
            article_column="A",
            image_column="B",
            start_row=2,
            desired_output_mb=5,
            cell_background_color="D9D9D9",
            local_images_by_source={
                "local_1": {"0010": local_image},
                "local_2": {},
            },
            source_order=["local_1", "server", "local_2"],
        )

        self.assertEqual(result["inserted"], 1)
        with zipfile.ZipFile(output) as archive:
            media_name = next(
                name for name in archive.namelist() if name.startswith("xl/media/")
            )
            with PILImage.open(BytesIO(archive.read(media_name))) as embedded:
                red, green, blue = embedded.convert("RGB").getpixel((20, 20))
                self.assertGreater(blue, red)
                self.assertGreater(blue, green)

    def test_endpoint_keeps_excel_in_memory_and_cleans_job_before_response(self):
        source = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "article"
        worksheet["B1"] = "image"
        worksheet["A2"] = "0010"
        workbook.save(source)
        workbook.close()
        source.seek(0)

        response = asyncio.run(
            server.process(
                DummyRequest(
                    {
                        "local_images": [],
                        "excel_filename": "Sensitive report.xlsx",
                        "excel_base64": base64.b64encode(source.read()).decode("ascii"),
                        "sheet_name": "Sheet",
                        "article_column": "A",
                        "image_column": "B",
                        "start_row": "2",
                        "desired_output_mb": "5",
                        "cell_background_color": "D9D9D9",
                        "source_order_json": '["server","local_1","local_2"]',
                        "local_image_manifest": "[]",
                    },
                    "test-progress",
                )
            )
        )
        response_body = asyncio.run(collect_streaming_response(response))

        self.assertEqual(response.status_code, 200)
        jobs_root = server.CONFIG["work_dir"] / "jobs"
        self.assertEqual(list(jobs_root.iterdir()), [])
        self.assertIn("Sensitive%20report_with%20Images.xlsx", response.headers["content-disposition"])
        self.assertEqual(response.headers["cache-control"], "no-store, private")
        with zipfile.ZipFile(BytesIO(response_body)) as archive:
            self.assertTrue(
                any(name.startswith("xl/media/") for name in archive.namelist())
            )
        public_log = server.CONFIG["public_log_file"].read_text(encoding="utf-8-sig")
        self.assertIn("Excel processing finished", public_log)
        self.assertNotIn("Sensitive report", public_log)
        progress = server.get_progress_snapshot("test-progress")
        self.assertEqual(progress["status"], "done")
        self.assertEqual(progress["percent"], 100)
        self.assertNotIn("Sensitive report", str(progress))


if __name__ == "__main__":
    unittest.main()
