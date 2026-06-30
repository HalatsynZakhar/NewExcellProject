from __future__ import annotations

import argparse
import asyncio
import base64
import binascii
import ctypes
import html
import io
import json
import os
import shutil
import stat
import sys
import tempfile
import threading
import time
import uuid
import xml.etree.ElementTree as ET
from contextlib import asynccontextmanager, contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse, urlsplit, urlunsplit

import requests
import uvicorn
from fastapi import HTTPException, Request, Response, UploadFile
from fastapi import FastAPI
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage
from starlette.formparsers import MultiPartParser


APP_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_FILE = APP_DIR / "config.json"
UTF8_BOM = b"\xef\xbb\xbf"
LOG_MUTEX_NAME = "Global\\ExcelImageServerSharedLog"

CONFIG: dict[str, Any] = {}
LOG_FILE: Path | None = None
MAX_LOG_LINES = 2000
INDEX_LOCK = threading.Lock()
INDEX_CACHE: dict[str, list[str]] = {}
INDEX_LOADED_AT = 0.0
INDEX_META: dict[str, Any] = {}
PROCESS_LOCK: threading.Lock | None = None
JOBS: dict[str, dict[str, Any]] = {}
PROGRESS_LOCK = threading.Lock()
PROGRESS_TTL_SECONDS = 60 * 60
CLIENT_HEARTBEAT_TIMEOUT_SECONDS = 20
SOURCE_IDS = ("server", "local_1", "local_2")
LOCAL_SOURCE_IDS = ("local_1", "local_2")
SUPPORTED_IMAGE_EXTENSIONS = (
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".bmp",
    ".tiff",
    ".webp",
)


class ProcessingCancelled(Exception):
    pass


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            try:
                reconfigure(encoding="utf-8", errors="backslashreplace")
            except (AttributeError, OSError, ValueError):
                pass


configure_console_encoding()


@contextmanager
def shared_log_lock():
    if os.name != "nt":
        yield
        return

    kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
    kernel32.CreateMutexW.argtypes = [
        ctypes.c_void_p,
        ctypes.c_bool,
        ctypes.c_wchar_p,
    ]
    kernel32.CreateMutexW.restype = ctypes.c_void_p
    kernel32.WaitForSingleObject.argtypes = [ctypes.c_void_p, ctypes.c_uint32]
    kernel32.WaitForSingleObject.restype = ctypes.c_uint32
    kernel32.ReleaseMutex.argtypes = [ctypes.c_void_p]
    kernel32.CloseHandle.argtypes = [ctypes.c_void_p]

    handle = kernel32.CreateMutexW(None, False, LOG_MUTEX_NAME)
    if not handle:
        raise OSError(ctypes.get_last_error(), "Could not create log mutex.")

    acquired = False
    try:
        result = kernel32.WaitForSingleObject(handle, 30_000)
        if result not in (0x00000000, 0x00000080):
            raise TimeoutError("Could not acquire log lock in 30 seconds.")
        acquired = True
        yield
    finally:
        if acquired:
            kernel32.ReleaseMutex(handle)
        kernel32.CloseHandle(handle)


def ensure_utf8_bom(path: Path) -> None:
    if not path.exists() or path.stat().st_size == 0:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(UTF8_BOM)
        return

    with path.open("rb") as file:
        prefix = file.read(len(UTF8_BOM))
    if prefix == UTF8_BOM:
        return

    path.write_bytes(UTF8_BOM + path.read_bytes())


def trim_log_file() -> None:
    if LOG_FILE is None or MAX_LOG_LINES < 1 or not LOG_FILE.exists():
        return

    lines = LOG_FILE.read_text(encoding="utf-8-sig").splitlines()
    if len(lines) <= MAX_LOG_LINES:
        return

    LOG_FILE.write_text(
        "\n".join(lines[-MAX_LOG_LINES:]) + "\n",
        encoding="utf-8-sig",
    )


def log(message: str) -> None:
    timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
    line = f"[{timestamp}] {message}"
    print(line, flush=True)

    if LOG_FILE is None:
        return

    try:
        with shared_log_lock():
            LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
            ensure_utf8_bom(LOG_FILE)
            with LOG_FILE.open("a", encoding="utf-8") as file:
                file.write(f"{line}\n")
            trim_log_file()
    except Exception as error:
        print(f"[{timestamp}] Could not write log {LOG_FILE}: {error}", flush=True)


def prepare_public_log() -> None:
    if LOG_FILE is None:
        return
    try:
        with shared_log_lock():
            LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
            ensure_utf8_bom(LOG_FILE)
            trim_log_file()
    except Exception:
        print("Could not prepare public log.", flush=True)


def path_is_inside(child: Path, parent: Path) -> bool:
    try:
        child.resolve(strict=False).relative_to(parent.resolve(strict=False))
        return True
    except ValueError:
        return False


def common_sharefiles_root(paths: list[Path]) -> Path | None:
    resolved = [path.resolve(strict=False) for path in paths]
    if not resolved:
        return None
    try:
        common = Path(os.path.commonpath([str(path) for path in resolved]))
    except ValueError:
        return None
    if common == common.anchor:
        return None
    if common.name.lower() == "public":
        return common.parent
    return None


def validate_private_work_dir(work_dir: Path, public_paths: list[Path]) -> None:
    work_dir = work_dir.resolve(strict=False)
    for public_path in public_paths:
        if path_is_inside(work_dir, public_path):
            raise ValueError(
                "work_dir must not be inside public XML, public log, or images folders."
            )

    sharefiles_root = common_sharefiles_root(public_paths)
    if sharefiles_root and path_is_inside(work_dir, sharefiles_root):
        raise ValueError("work_dir must not be inside the public shared file tree.")


def normalize_base_url(value: str) -> str:
    parsed = urlsplit(str(value).strip())
    path_parts = [part for part in parsed.path.split("/") if part]
    normalized_path = "/" + "/".join(path_parts) if path_parts else ""
    normalized_path = normalized_path.rstrip("/") + "/"
    return urlunsplit(
        (
            parsed.scheme,
            parsed.netloc,
            normalized_path,
            "",
            "",
        )
    )


def load_config(config_file: Path = DEFAULT_CONFIG_FILE) -> dict[str, Any]:
    if not config_file.exists():
        raise FileNotFoundError(f"Config file not found: {config_file}")

    raw = json.loads(config_file.read_text(encoding="utf-8"))
    required = ("images_dir", "images_base_url")
    missing = [key for key in required if not raw.get(key)]
    if missing:
        raise ValueError("Missing config keys: " + ", ".join(missing))
    if not raw.get("xml_path") and not raw.get("xml_url"):
        raise ValueError("Either xml_path or xml_url must be configured.")

    public_log_dir = Path(raw.get("public_log_dir") or (APP_DIR / "logs")).expanduser()
    public_log_filename = raw.get("public_log_filename", "excel_image_server.log")
    if Path(public_log_filename).name != public_log_filename:
        raise ValueError("public_log_filename must be a filename, not a path.")

    work_dir = Path(raw.get("work_dir", APP_DIR / "work")).expanduser()
    xml_path = Path(raw["xml_path"]).expanduser() if raw.get("xml_path") else None
    images_dir = Path(raw["images_dir"]).expanduser()
    public_paths = [public_log_dir, images_dir]
    if xml_path is not None:
        public_paths.append(xml_path.parent)
    validate_private_work_dir(work_dir, public_paths)
    max_upload_mb = int(raw.get("max_upload_mb", 100))
    max_local_images_upload_mb = int(
        raw.get("max_local_images_upload_mb", 500)
    )
    hard_max_output_mb = min(100, int(raw.get("hard_max_output_mb", 100)))
    default_desired_output_mb = int(float(raw.get("default_desired_output_mb", 20)))
    max_parallel_jobs = int(raw.get("max_parallel_jobs", 1))
    if max_upload_mb < 1:
        raise ValueError("max_upload_mb must be positive.")
    if max_local_images_upload_mb < 1:
        raise ValueError("max_local_images_upload_mb must be positive.")
    if hard_max_output_mb < 1:
        raise ValueError("hard_max_output_mb must be positive.")
    if not 1 <= default_desired_output_mb <= 100:
        raise ValueError("default_desired_output_mb must be from 1 to 100.")
    if max_parallel_jobs < 1:
        raise ValueError("max_parallel_jobs must be positive.")

    return {
        "host": str(raw.get("host", "0.0.0.0")),
        "port": int(raw.get("port", 8091)),
        "xml_path": xml_path,
        "xml_url": str(raw.get("xml_url", "")),
        "images_dir": images_dir,
        "images_base_url": normalize_base_url(str(raw["images_base_url"])),
        "work_dir": work_dir,
        "public_log_file": public_log_dir / public_log_filename,
        "public_log_url": str(
            raw.get(
                "public_log_url",
                "/xml/excel_image_server.log",
            )
        ),
        "max_log_lines": int(raw.get("max_log_lines", 2000)),
        "max_upload_mb": max_upload_mb,
        "max_local_images_upload_mb": max_local_images_upload_mb,
        "hard_max_output_mb": hard_max_output_mb,
        "default_desired_output_mb": default_desired_output_mb,
        "max_parallel_jobs": 1,
        "job_ttl_minutes": int(raw.get("job_ttl_minutes", 120)),
        "xml_cache_seconds": int(raw.get("xml_cache_seconds", 300)),
        "default_start_row": int(raw.get("default_start_row", 2)),
        "default_article_column": str(raw.get("default_article_column", "A")),
        "default_image_column": str(raw.get("default_image_column", "B")),
        "target_image_width_px": int(raw.get("target_image_width_px", 300)),
        "adjust_row_height": bool(raw.get("adjust_row_height", True)),
        "default_cell_background_color": str(
            raw.get("default_cell_background_color", "D9D9D9")
        ),
        "image_padding_px": int(raw.get("image_padding_px", 1)),
        "image_width_guard_px": int(raw.get("image_width_guard_px", 6)),
        "download_timeout_seconds": int(raw.get("download_timeout_seconds", 20)),
    }


def configure_runtime(config: dict[str, Any]) -> None:
    global CONFIG, LOG_FILE, MAX_LOG_LINES, PROCESS_LOCK
    CONFIG = config
    LOG_FILE = config["public_log_file"]
    MAX_LOG_LINES = config["max_log_lines"]
    PROCESS_LOCK = threading.Lock()
    with PROGRESS_LOCK:
        JOBS.clear()
    prepare_public_log()
    config["work_dir"].mkdir(parents=True, exist_ok=True)
    secure_delete_tree(config["work_dir"] / "jobs")
    (config["work_dir"] / "jobs").mkdir(parents=True, exist_ok=True)
    configure_private_temp_dir(config)
    for legacy_name in ("uploads", "results", "images"):
        secure_delete_tree(config["work_dir"] / legacy_name)


def configure_private_temp_dir(config: dict[str, Any]) -> None:
    temp_dir = config["work_dir"] / "python_temp"
    secure_delete_tree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    config["python_temp_dir"] = temp_dir
    MultiPartParser.spool_max_size = config["max_upload_mb"] * 1024 * 1024
    apply_private_temp_dir(config)


def apply_private_temp_dir(config: dict[str, Any]) -> None:
    temp_dir = config["python_temp_dir"]
    temp_text = str(temp_dir)
    os.environ["TMP"] = temp_text
    os.environ["TEMP"] = temp_text
    os.environ["TMPDIR"] = temp_text
    tempfile.tempdir = temp_text


def ensure_private_temp_dir() -> None:
    work_dir = CONFIG["work_dir"]
    work_dir.mkdir(parents=True, exist_ok=True)
    (work_dir / "jobs").mkdir(parents=True, exist_ok=True)
    temp_dir = CONFIG.get("python_temp_dir") or (work_dir / "python_temp")
    temp_dir = Path(temp_dir)
    CONFIG["python_temp_dir"] = temp_dir
    if temp_dir.exists() and not temp_dir.is_dir():
        secure_delete_tree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    apply_private_temp_dir(CONFIG)


def normalize_article(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_article_key(value: Any) -> str:
    text = normalize_article(value)
    normalized = []
    for character in text:
        if character.isalnum() or character == " ":
            normalized.append(character)
        else:
            normalized.append("-")
    return "".join(normalized).casefold()


def normalize_local_article(value: Any, *, for_excel: bool) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    normalized = []
    for character in text:
        if character.isalnum() or character == " ":
            normalized.append(character)
        elif not for_excel and character == "_":
            normalized.append(character)
        else:
            normalized.append("-")
    return "".join(normalized).lower()


def validate_source_order(value: str | list[str] | None) -> list[str]:
    try:
        parsed = json.loads(value) if isinstance(value, str) else value
    except json.JSONDecodeError as error:
        raise HTTPException(status_code=400, detail="Некоректний порядок джерел.") from error
    if parsed is None:
        parsed = list(SOURCE_IDS)
    if not isinstance(parsed, list):
        raise HTTPException(status_code=400, detail="Некоректний порядок джерел.")
    order = [str(item) for item in parsed]
    if any(source not in SOURCE_IDS for source in order) or len(set(order)) != len(order):
        raise HTTPException(status_code=400, detail="Пріоритети джерел мають бути унікальними.")
    return order


def normalize_progress_id(value: str | None) -> str:
    text = (value or "").strip()
    if not text or len(text) > 80:
        return ""
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_")
    if any(character not in allowed for character in text):
        return ""
    return text


def cleanup_old_progress() -> None:
    cutoff = time.monotonic() - PROGRESS_TTL_SECONDS
    with PROGRESS_LOCK:
        expired = [
            progress_id
            for progress_id, item in JOBS.items()
            if float(item.get("updated_monotonic", 0)) < cutoff
        ]
        for progress_id in expired:
            JOBS.pop(progress_id, None)


def set_progress(
    progress_id: str,
    percent: float,
    message: str,
    *,
    status: str = "running",
    inserted: int | None = None,
    total: int | None = None,
    report: dict[str, Any] | None = None,
) -> None:
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return
    with PROGRESS_LOCK:
        existing = JOBS.get(progress_id, {})
        item: dict[str, Any] = {
            "ok": True,
            "status": status,
            "percent": max(0, min(100, round(float(percent), 1))),
            "message": message,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "updated_monotonic": time.monotonic(),
            "created_monotonic": existing.get("created_monotonic", time.monotonic()),
        }
        for key in (
            "cancel_requested",
            "heartbeat_required",
            "last_client_seen_monotonic",
            "last_client_seen_at",
        ):
            if key in existing:
                item[key] = existing[key]
        if inserted is not None:
            item["inserted"] = inserted
        elif "inserted" in existing:
            item["inserted"] = existing["inserted"]
        if total is not None:
            item["total"] = total
        elif "total" in existing:
            item["total"] = existing["total"]
        if report is not None:
            item["report"] = report
        elif "report" in existing:
            item["report"] = existing["report"]
        JOBS[progress_id] = item


def request_cancel(progress_id: str) -> bool:
    cleanup_old_progress()
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return False
    with PROGRESS_LOCK:
        item = JOBS.get(progress_id)
        if not item:
            return False
        status = str(item.get("status", ""))
        if status in {"done", "error", "cancelled"}:
            return False
        item["cancel_requested"] = True
        item["status"] = "cancelling"
        item["message"] = "Скасування обробки..."
        item["updated_at"] = datetime.now().isoformat(timespec="seconds")
        item["updated_monotonic"] = time.monotonic()
    return True


def mark_progress_requires_heartbeat(progress_id: str) -> None:
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return
    now = time.monotonic()
    with PROGRESS_LOCK:
        item = JOBS.setdefault(
            progress_id,
            {
                "ok": True,
                "status": "running",
                "percent": 0,
                "message": "Очікування початку обробки.",
                "created_monotonic": now,
                "updated_monotonic": now,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
        item["heartbeat_required"] = True
        item["last_client_seen_monotonic"] = now
        item["last_client_seen_at"] = datetime.now().isoformat(timespec="seconds")
        item.setdefault("created_monotonic", now)


def mark_client_heartbeat(progress_id: str) -> bool:
    cleanup_old_progress()
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return False
    now = time.monotonic()
    with PROGRESS_LOCK:
        item = JOBS.setdefault(
            progress_id,
            {
                "ok": True,
                "status": "waiting",
                "percent": 0,
                "message": "Очікування початку обробки.",
                "created_monotonic": now,
                "updated_monotonic": now,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
        status = str(item.get("status", ""))
        if status in {"done", "error", "cancelled"}:
            return False
        item["last_client_seen_monotonic"] = now
        item["last_client_seen_at"] = datetime.now().isoformat(timespec="seconds")
        item.setdefault("created_monotonic", now)
        return True


def cancel_if_client_stale(progress_id: str) -> bool:
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return False
    now = time.monotonic()
    with PROGRESS_LOCK:
        item = JOBS.get(progress_id)
        if not item or not item.get("heartbeat_required"):
            return False
        status = str(item.get("status", ""))
        if status in {"done", "error", "cancelled"}:
            return False
        last_seen = item.get("last_client_seen_monotonic", item.get("created_monotonic", now))
        if now - float(last_seen) <= CLIENT_HEARTBEAT_TIMEOUT_SECONDS:
            return False
        item["cancel_requested"] = True
        item["status"] = "cancelling"
        item["message"] = "Вкладку закрито. Обробку скасовано."
        item["updated_at"] = datetime.now().isoformat(timespec="seconds")
        item["updated_monotonic"] = now
        return True


def is_cancel_requested(progress_id: str) -> bool:
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return False
    if cancel_if_client_stale(progress_id):
        return True
    with PROGRESS_LOCK:
        return bool(JOBS.get(progress_id, {}).get("cancel_requested"))


def raise_if_cancelled(progress_id: str) -> None:
    if is_cancel_requested(progress_id):
        raise ProcessingCancelled("Обробку скасовано користувачем.")


def get_progress_snapshot(progress_id: str) -> dict[str, Any]:
    cleanup_old_progress()
    progress_id = normalize_progress_id(progress_id)
    if not progress_id:
        return {
            "ok": False,
            "status": "unknown",
            "percent": 0,
            "message": "Невідомий запит.",
        }
    with PROGRESS_LOCK:
        item = dict(JOBS.get(progress_id, {}))
    if not item:
        return {
            "ok": False,
            "status": "unknown",
            "percent": 0,
            "message": "Очікування початку обробки.",
        }
    item.pop("updated_monotonic", None)
    return item


def load_xml_index(force: bool = False) -> dict[str, list[str]]:
    global INDEX_CACHE, INDEX_LOADED_AT, INDEX_META
    now = time.monotonic()
    with INDEX_LOCK:
        if (
            not force
            and INDEX_CACHE
            and now - INDEX_LOADED_AT < CONFIG["xml_cache_seconds"]
        ):
            return INDEX_CACHE

        xml_content: bytes | None = None
        source = ""
        xml_path = CONFIG.get("xml_path")
        if xml_path and Path(xml_path).is_file():
            source = str(xml_path)
            xml_content = Path(xml_path).read_bytes()
        elif CONFIG.get("xml_url"):
            source = CONFIG["xml_url"]
            response = requests.get(
                CONFIG["xml_url"],
                timeout=CONFIG["download_timeout_seconds"],
            )
            response.raise_for_status()
            xml_content = response.content
        else:
            raise FileNotFoundError("No available XML source configured.")

        root = ET.fromstring(xml_content)

        index: dict[str, list[str]] = {}
        image_count = 0
        for product in root.findall(".//product"):
            article = product.attrib.get("article", "")
            key = normalize_article_key(article)
            if not key:
                continue
            urls = [
                (image.text or "").strip()
                for image in product.findall("image")
                if (image.text or "").strip()
            ]
            if not urls:
                continue
            index[key] = urls
            image_count += len(urls)

        INDEX_CACHE = index
        INDEX_LOADED_AT = now
        INDEX_META = {
            "loaded_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "products_count": len(index),
            "images_count": image_count,
            "source_generated_at": root.attrib.get("generated_at", ""),
            "source": source,
        }
        log(
            "XML index loaded: "
            f"products={len(index)}, images={image_count}"
        )
        return INDEX_CACHE


def image_url_to_local_path(image_url: str) -> Path | None:
    base_url = CONFIG["images_base_url"]
    if not image_url.lower().startswith(base_url.lower()):
        return None

    relative_url = image_url[len(base_url) :]
    relative_path = unquote(urlparse(relative_url).path).lstrip("/")
    parts = [part for part in relative_path.split("/") if part not in ("", ".", "..")]
    if not parts:
        return None

    root = CONFIG["images_dir"].resolve(strict=False)
    candidate = root.joinpath(*parts).resolve(strict=False)
    try:
        candidate.relative_to(root)
    except ValueError:
        return None
    return candidate


def column_to_index(value: str) -> int:
    text = value.strip()
    if not text:
        raise ValueError("Стовпець не може бути порожнім.")
    if text.isdigit():
        number = int(text)
        if number < 1:
            raise ValueError("Номер стовпця має бути додатним.")
        return number
    return column_index_from_string(text.upper())


def reset_stream(file_source: Any) -> None:
    seek = getattr(file_source, "seek", None)
    if callable(seek):
        seek(0)


def workbook_suffix(file_source: Any, suffix: str | None = None) -> str:
    if suffix:
        return suffix.lower()
    if isinstance(file_source, (str, os.PathLike)):
        return Path(file_source).suffix.lower()
    return ""


def workbook_size_bytes(file_source: Any, size_bytes: int | None = None) -> int:
    if size_bytes is not None:
        return size_bytes
    if isinstance(file_source, (str, os.PathLike)):
        return Path(file_source).stat().st_size
    getbuffer = getattr(file_source, "getbuffer", None)
    if callable(getbuffer):
        return len(getbuffer())
    tell = getattr(file_source, "tell", None)
    seek = getattr(file_source, "seek", None)
    if callable(tell) and callable(seek):
        current = tell()
        seek(0, os.SEEK_END)
        size = tell()
        seek(current)
        return size
    return 0


def load_workbook_safely(
    file_source: Any,
    *,
    read_only: bool = False,
    keep_vba: bool = False,
    data_only: bool = False,
) -> Any:
    reset_stream(file_source)
    return load_workbook(
        file_source,
        read_only=read_only,
        keep_vba=keep_vba,
        data_only=data_only,
        keep_links=True,
    )


def get_sheet_names(file_source: Any, suffix: str | None = None) -> list[str]:
    suffix = workbook_suffix(file_source, suffix)
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError("Підтримуються лише .xlsx та .xlsm.")
    workbook = load_workbook_safely(
        file_source,
        read_only=True,
        keep_vba=suffix == ".xlsm",
    )
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def inspect_workbook(
    file_source: Any,
    preview_rows: int = 12,
    preview_cols: int = 12,
    suffix: str | None = None,
) -> dict[str, Any]:
    suffix = workbook_suffix(file_source, suffix)
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError("Підтримуються лише .xlsx та .xlsm.")

    workbook = load_workbook_safely(
        file_source,
        read_only=True,
        keep_vba=suffix == ".xlsm",
        data_only=False,
    )
    try:
        previews: dict[str, dict[str, Any]] = {}
        for worksheet in workbook.worksheets:
            max_column = min(max(worksheet.max_column or 1, 1), preview_cols)
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(max(worksheet.max_row or 1, 1), preview_rows),
                max_col=max_column,
                values_only=True,
            ):
                rows.append(["" if value is None else str(value) for value in row])
            previews[worksheet.title] = {
                "rows": rows,
                "max_row": worksheet.max_row or 0,
                "max_column": worksheet.max_column or 0,
            }
        return {"sheets": list(workbook.sheetnames), "previews": previews}
    finally:
        workbook.close()


def inspect_article_requirements(
    file_source: Any,
    sheet_name: str,
    article_column: str,
    start_row: int,
    suffix: str | None = None,
) -> dict[str, Any]:
    suffix = workbook_suffix(file_source, suffix)
    workbook = load_workbook_safely(
        file_source,
        read_only=True,
        keep_vba=suffix == ".xlsm",
        data_only=False,
    )
    try:
        selected_sheet = sheet_name.strip() or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(f"Аркуш не знайдено: {selected_sheet}")
        worksheet = workbook[selected_sheet]
        article_col_idx = column_to_index(article_column)
        start_row = max(1, int(start_row))
        index = load_xml_index()
        requirements: list[dict[str, Any]] = []
        seen_keys: set[str] = set()
        for row in range(start_row, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=article_col_idx).value
            local_key = normalize_local_article(value, for_excel=True)
            if not local_key or local_key in seen_keys:
                continue
            seen_keys.add(local_key)
            requirements.append(
                {
                    "key": local_key,
                    "server_available": bool(index.get(normalize_article_key(value))),
                }
            )
        return {
            "sheet": selected_sheet,
            "requirements": requirements,
        }
    finally:
        workbook.close()


def read_excel_base64_to_memory(
    encoded_excel: str,
    original_filename: str | None,
) -> tuple[io.BytesIO, str, int]:
    filename = Path(original_filename or "input.xlsx").name
    suffix = Path(filename).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        raise HTTPException(
            status_code=400,
            detail="Підтримуються лише .xlsx та .xlsm. Старий .xls поки не обробляється.",
        )

    try:
        data = base64.b64decode(encoded_excel, validate=True)
    except (binascii.Error, ValueError) as error:
        raise HTTPException(status_code=400, detail="Excel-файл передано некоректно.") from error

    max_bytes = CONFIG["max_upload_mb"] * 1024 * 1024
    if len(data) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Файл більший за ліміт {CONFIG['max_upload_mb']} МБ.",
        )
    return io.BytesIO(data), suffix, len(data)


def read_excel_form_to_memory(form_data: Any) -> tuple[io.BytesIO, str, int, str | None]:
    encoded_excel = form_data.get("excel_base64")
    original_filename = form_data.get("excel_filename")
    if isinstance(encoded_excel, str) and encoded_excel:
        buffer, suffix, size = read_excel_base64_to_memory(
            encoded_excel,
            str(original_filename or "input.xlsx"),
        )
        return buffer, suffix, size, str(original_filename or "input.xlsx")

    raise HTTPException(status_code=400, detail="Excel-файл не передано.")


def close_upload(uploaded_file: UploadFile) -> None:
    try:
        uploaded_file.file.close()
    except Exception:
        pass


def save_local_image_uploads(
    uploaded_files: list[UploadFile],
    manifest_json: str,
    job_dir: Path,
) -> dict[str, dict[str, Path]]:
    result: dict[str, dict[str, Path]] = {
        source: {} for source in LOCAL_SOURCE_IDS
    }
    if not uploaded_files:
        return result
    try:
        manifest = json.loads(manifest_json or "[]")
    except json.JSONDecodeError as error:
        raise HTTPException(status_code=400, detail="Некоректний опис локальних зображень.") from error
    if not isinstance(manifest, list) or len(manifest) != len(uploaded_files):
        raise HTTPException(status_code=400, detail="Локальні зображення передані не повністю.")

    local_dir = job_dir / "local"
    local_dir.mkdir(parents=True, exist_ok=True)
    max_total_bytes = int(CONFIG.get("max_local_images_upload_mb", 500)) * 1024 * 1024
    total_bytes = 0

    for index, (uploaded_file, item) in enumerate(zip(uploaded_files, manifest)):
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="Некоректний опис локального зображення.")
        source = str(item.get("source", ""))
        key = str(item.get("key", ""))
        if source not in LOCAL_SOURCE_IDS or not key:
            raise HTTPException(status_code=400, detail="Некоректне джерело локального зображення.")

        original_name = (uploaded_file.filename or "").replace("\\", "/").rsplit("/", 1)[-1]
        suffix = Path(original_name).suffix.lower()
        original_stem = Path(original_name).stem
        if suffix not in SUPPORTED_IMAGE_EXTENSIONS:
            raise HTTPException(status_code=400, detail="Непідтримуваний формат локального зображення.")
        if normalize_local_article(original_stem, for_excel=False) != key:
            raise HTTPException(status_code=400, detail="Назва локального зображення не відповідає артикулу.")

        destination = local_dir / f"image_{index}{suffix}"
        with destination.open("wb") as output:
            while True:
                chunk = uploaded_file.file.read(1024 * 1024)
                if not chunk:
                    break
                total_bytes += len(chunk)
                if total_bytes > max_total_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail=(
                            "Загальний розмір локальних зображень перевищує "
                            f"{CONFIG.get('max_local_images_upload_mb', 500)} МБ."
                        ),
                    )
                output.write(chunk)
        try:
            with PILImage.open(destination) as image:
                image.verify()
        except Exception as error:
            raise HTTPException(status_code=400, detail="Один із локальних файлів не є зображенням.") from error

        if key not in result[source]:
            result[source][key] = destination
        else:
            wipe_file(destination)

    return result


def make_download_filename(original_filename: str | None, suffix: str) -> str:
    original_name = (original_filename or "").replace("\\", "/").rsplit("/", 1)[-1]
    stem = Path(original_name).stem.strip()
    stem = "".join(
        "_" if character in '<>:"/\\|?*' or ord(character) < 32 else character
        for character in stem
    ).rstrip(" .")
    if not stem:
        stem = "Excel"
    return f"{stem}_with Images{suffix}"


def quote_header_filename(filename: str) -> str:
    return quote(filename, safe="")


def sanitize_hex_color(value: str, default: str = "D9D9D9") -> str:
    text = (value or "").strip().lstrip("#").upper()
    if len(text) == 8:
        text = text[-6:]
    if len(text) == 6 and all(character in "0123456789ABCDEF" for character in text):
        return text
    return default


def column_width_to_pixels(worksheet: Any, column_letter: str) -> int:
    width = worksheet.column_dimensions[column_letter].width
    if width is None:
        width = worksheet.sheet_format.defaultColWidth or 8.43
    if width < 1:
        return max(12, int(width * 12))
    return max(12, int(width * 7 + 5))


def encode_image_for_budget(
    source_image: PILImage.Image,
    target_bytes: int,
) -> tuple[bytes, int, int]:
    if "A" in source_image.getbands() or "transparency" in source_image.info:
        rgba_image = source_image.convert("RGBA")
        image = PILImage.new("RGB", source_image.size, (255, 255, 255))
        image.paste(rgba_image, mask=rgba_image.getchannel("A"))
    else:
        image = source_image.convert("RGB")

    width, height = image.size
    target_bytes = max(512, int(target_bytes))
    maximum_quality = 95

    def encode(current_image: PILImage.Image, current_quality: int) -> bytes:
        buffer = io.BytesIO()
        current_image.save(
            buffer,
            format="JPEG",
            quality=current_quality,
            optimize=True,
        )
        return buffer.getvalue()

    def best_encoded_for_size(current_image: PILImage.Image) -> bytes:
        high_quality_data = encode(current_image, maximum_quality)
        if len(high_quality_data) <= target_bytes:
            return high_quality_data

        low, high = 1, maximum_quality - 1
        best_fit: bytes | None = None
        smallest_data = high_quality_data
        while low <= high:
            current_quality = (low + high) // 2
            data = encode(current_image, current_quality)
            if len(data) < len(smallest_data):
                smallest_data = data
            if len(data) <= target_bytes:
                best_fit = data
                low = current_quality + 1
            else:
                high = current_quality - 1
        return best_fit or smallest_data

    encoded = best_encoded_for_size(image)
    if len(encoded) <= target_bytes:
        return encoded, width, height

    best_data = encoded
    best_width, best_height = width, height
    scale = 0.85
    while min(width, height) * scale >= 48:
        resized_width = max(48, int(width * scale))
        resized_height = max(48, int(height * scale))
        resized = image.resize((resized_width, resized_height), PILImage.Resampling.LANCZOS)
        data = best_encoded_for_size(resized)
        if len(data) < len(best_data):
            best_data = data
            best_width, best_height = resized_width, resized_height
        if len(data) <= target_bytes:
            return data, resized_width, resized_height
        scale *= 0.85

    return best_data, best_width, best_height


def prepare_image(
    image_source: str | Path,
    job_images_dir: Path,
    target_bytes_per_image: int,
) -> tuple[Path, int, int, str, int]:
    image_bytes: bytes | None = None
    if isinstance(image_source, Path):
        source = "local"
        image_bytes = image_source.read_bytes()
    else:
        source = image_source
        local_path = image_url_to_local_path(image_source)
        if local_path and local_path.is_file():
            source = str(local_path)
            image_bytes = local_path.read_bytes()
        else:
            response = requests.get(
                image_source,
                timeout=CONFIG["download_timeout_seconds"],
            )
            response.raise_for_status()
            image_bytes = response.content

    with PILImage.open(io.BytesIO(image_bytes)) as source_image:
        source_format = (source_image.format or "").upper()
        width, height = source_image.size
        preserved_formats = {
            "JPEG": ".jpg",
            "JPG": ".jpg",
            "PNG": ".png",
            "GIF": ".gif",
        }
        if (
            len(image_bytes) <= target_bytes_per_image
            and source_format in preserved_formats
        ):
            encoded = image_bytes
            suffix = preserved_formats[source_format]
        else:
            encoded, width, height = encode_image_for_budget(
                source_image,
                target_bytes_per_image,
            )
            suffix = ".jpg"

    output_path = job_images_dir / f"{uuid.uuid4().hex}{suffix}"
    output_path.write_bytes(encoded)

    return output_path, width, height, source, len(encoded)


def process_excel(
    input_path: Any,
    output_path: Any,
    sheet_name: str,
    article_column: str,
    image_column: str,
    start_row: int,
    desired_output_mb: float,
    cell_background_color: str,
    use_cell_background: bool = True,
    local_images_by_source: dict[str, dict[str, Path]] | None = None,
    source_order: list[str] | None = None,
    progress_callback: Any | None = None,
    cancel_callback: Any | None = None,
    input_suffix: str | None = None,
    input_size_bytes: int | None = None,
    temp_dir: Path | None = None,
) -> dict[str, Any]:
    def report(percent: float, message: str, **extra: Any) -> None:
        if progress_callback is not None:
            progress_callback(percent, message, **extra)

    def check_cancelled() -> None:
        if cancel_callback is not None:
            cancel_callback()

    suffix = workbook_suffix(input_path, input_suffix)
    keep_vba = suffix == ".xlsm"
    report(10, "Відкриття Excel-книги...")
    workbook = load_workbook_safely(input_path, keep_vba=keep_vba)
    temp_images: list[Path] = []
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Аркуш не знайдено: {sheet_name}")

        worksheet = workbook[sheet_name]
        article_col_idx = column_to_index(article_column)
        image_col_idx = column_to_index(image_column)
        image_col_letter = get_column_letter(image_col_idx)
        start_row = max(1, int(start_row))
        hard_max_output_mb = CONFIG["hard_max_output_mb"]
        desired_output_mb = min(100, max(1, int(float(desired_output_mb))))
        desired_output_mb = min(desired_output_mb, hard_max_output_mb)
        background_color = sanitize_hex_color(
            cell_background_color,
            CONFIG["default_cell_background_color"],
        ) if use_cell_background else ""
        cell_fill = (
            PatternFill(fill_type="solid", fgColor=background_color)
            if use_cell_background
            else None
        )
        padding_px = max(0, int(CONFIG.get("image_padding_px", 1)))
        width_guard_px = max(0, int(CONFIG.get("image_width_guard_px", 6)))
        local_images_by_source = local_images_by_source or {
            source: {} for source in LOCAL_SOURCE_IDS
        }
        source_order = validate_source_order(source_order)

        index = load_xml_index()
        if temp_dir is not None:
            job_images_dir = temp_dir / "images"
        elif isinstance(output_path, (str, os.PathLike)):
            job_images_dir = Path(output_path).parent / "images"
        else:
            job_images_dir = CONFIG["work_dir"] / "jobs" / f"images_{uuid.uuid4().hex}"
        job_images_dir.mkdir(parents=True, exist_ok=True)

        report(18, "Пошук артикулів і доступних зображень...")
        planned_rows: list[tuple[int, str, str | Path]] = []
        rows_seen = 0
        not_found: list[str] = []
        max_row = worksheet.max_row or start_row
        total_scan_rows = max(1, max_row - start_row + 1)
        scan_step = max(1, total_scan_rows // 20)
        for row in range(start_row, max_row + 1):
            check_cancelled()
            raw_article = worksheet.cell(row=row, column=article_col_idx).value
            article = "" if raw_article is None else str(raw_article).strip()
            if not article:
                continue
            rows_seen += 1
            image_urls = index.get(normalize_article_key(raw_article), [])
            local_key = normalize_local_article(raw_article, for_excel=True)
            selected_source: str | Path | None = None
            for source_id in source_order:
                if source_id == "server" and image_urls:
                    selected_source = image_urls[0]
                    break
                local_path = local_images_by_source.get(source_id, {}).get(local_key)
                if local_path is not None:
                    selected_source = local_path
                    break
            if selected_source is None:
                not_found.append(article)
                continue
            planned_rows.append((row, article, selected_source))
            if (row - start_row + 1) % scan_step == 0:
                scan_percent = 18 + ((row - start_row + 1) / total_scan_rows) * 12
                report(scan_percent, "Пошук артикулів і доступних зображень...")

        image_count = len(planned_rows)
        report(30, "Підготовка вставки зображень...", total=image_count)
        desired_bytes = int(desired_output_mb * 1024 * 1024)
        workbook_bytes = workbook_size_bytes(input_path, input_size_bytes)
        estimated_excel_overhead = image_count * 2600
        available_image_bytes = max(0, desired_bytes - workbook_bytes - estimated_excel_overhead)
        target_bytes_per_image = max(
            512,
            int(available_image_bytes / max(1, image_count)),
        )
        column_pixels = column_width_to_pixels(worksheet, image_col_letter)
        image_box_size = max(24, column_pixels - padding_px - width_guard_px)

        inserted = 0
        failed: list[str] = []
        sources: list[dict[str, str]] = []

        for index, (row, article, image_source) in enumerate(planned_rows, start=1):
            check_cancelled()
            try:
                if image_count:
                    report(
                        ((index - 1) / image_count) * 100,
                        "Вставка зображень у Excel...",
                        inserted=inserted,
                        total=image_count,
                    )
                if cell_fill is not None:
                    worksheet.cell(row=row, column=image_col_idx).fill = cell_fill
                image_path, embedded_width, embedded_height, source, image_size = prepare_image(
                    image_source,
                    job_images_dir,
                    target_bytes_per_image,
                )
                temp_images.append(image_path)
                excel_image = OpenpyxlImage(str(image_path))
                display_scale = min(
                    image_box_size / max(1, embedded_width),
                    image_box_size / max(1, embedded_height),
                )
                display_width = max(1, int(embedded_width * display_scale))
                display_height = max(1, int(embedded_height * display_scale))
                excel_image.width = display_width
                excel_image.height = display_height
                left_offset_px = padding_px
                top_offset_px = padding_px
                marker = AnchorMarker(
                    col=image_col_idx - 1,
                    colOff=pixels_to_EMU(left_offset_px),
                    row=row - 1,
                    rowOff=pixels_to_EMU(top_offset_px),
                )
                excel_image.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(
                        pixels_to_EMU(display_width),
                        pixels_to_EMU(display_height),
                    ),
                )
                worksheet.add_image(excel_image)
                if CONFIG["adjust_row_height"]:
                    worksheet.row_dimensions[row].height = (
                        display_height + padding_px + 1
                    ) * 0.75
                inserted += 1
                sources.append(
                    {
                        "article": article,
                        "source": source,
                        "image_size_kb": str(round(image_size / 1024, 1)),
                    }
                )
            except Exception as error:
                failed.append(f"{article}: {error}")
                log("Image insert failed in current job.")
            if image_count:
                report(
                    (index / image_count) * 100,
                    "Вставка зображень у Excel...",
                    inserted=inserted,
                    total=image_count,
                )

        report(92, "Збереження Excel-файлу...", inserted=inserted, total=image_count)
        check_cancelled()
        reset_stream(output_path)
        truncate = getattr(output_path, "truncate", None)
        if callable(truncate):
            truncate(0)
        workbook.save(output_path)
        output_size_bytes = workbook_size_bytes(output_path)
        output_size_mb = output_size_bytes / (1024 * 1024)
        if output_size_mb > hard_max_output_mb:
            if isinstance(output_path, (str, os.PathLike)):
                wipe_file(Path(output_path))
            else:
                output_path.seek(0)
                output_path.truncate(0)
            raise ValueError(
                f"Файл результату має розмір {output_size_mb:.1f} МБ, це більше жорсткого ліміту "
                f"{hard_max_output_mb} МБ."
            )
        reset_stream(output_path)
        return {
            "rows_seen": rows_seen,
            "inserted": inserted,
            "not_found": not_found,
            "failed": failed,
            "sources": sources[:50],
            "output_file": str(output_path) if isinstance(output_path, (str, os.PathLike)) else "",
            "output_size_mb": round(output_size_mb, 2),
            "desired_output_mb": desired_output_mb,
            "target_bytes_per_image": target_bytes_per_image,
            "image_column_width_px": column_pixels,
            "cell_background_color": background_color,
            "use_cell_background": use_cell_background,
        }
    finally:
        workbook.close()
        # Openpyxl reads image files during save; remove them only after saving.
        for path in temp_images:
            wipe_file(path)


def cleanup_old_jobs() -> None:
    ttl = timedelta(minutes=CONFIG["job_ttl_minutes"])
    cutoff = datetime.now() - ttl
    root = CONFIG["work_dir"] / "jobs"
    if not root.exists():
        return
    for child in root.iterdir():
        if not child.is_dir():
            continue
        try:
            if not any(child.iterdir()):
                secure_delete_tree(child)
                continue
            modified = datetime.fromtimestamp(child.stat().st_mtime)
            if modified < cutoff:
                secure_delete_tree(child)
        except OSError:
            pass


def wipe_file(path: Path) -> None:
    try:
        if not path.is_file():
            return
        size = path.stat().st_size
        with path.open("r+b", buffering=0) as file:
            chunk = b"\0" * min(1024 * 1024, max(size, 1))
            remaining = size
            while remaining > 0:
                current = min(len(chunk), remaining)
                file.write(chunk[:current])
                remaining -= current
            file.flush()
            os.fsync(file.fileno())
        path.unlink(missing_ok=True)
    except OSError:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def force_remove_error(func: Any, path: str, _exc_info: Any) -> None:
    try:
        os.chmod(path, stat.S_IWRITE)
        func(path)
    except OSError:
        pass


def secure_delete_tree(path: Path) -> bool:
    if not path.exists():
        return True
    if path.is_file():
        wipe_file(path)
        return not path.exists()
    for attempt in range(5):
        for child in sorted(path.rglob("*"), key=lambda item: len(item.parts), reverse=True):
            try:
                if child.is_file():
                    wipe_file(child)
                elif child.is_dir():
                    child.rmdir()
            except OSError:
                pass
        try:
            path.rmdir()
        except OSError:
            try:
                shutil.rmtree(path, ignore_errors=False, onerror=force_remove_error)
            except OSError:
                pass
        if not path.exists():
            return True
        time.sleep(0.05 * (attempt + 1))
    return not path.exists()


def make_job_dir(kind: str = "job") -> Path:
    ensure_private_temp_dir()
    cleanup_old_jobs()
    job_id = uuid.uuid4().hex
    job_dir = CONFIG["work_dir"] / "jobs" / f"{kind}_{job_id}"
    job_dir.mkdir(parents=True, exist_ok=False)
    return job_dir


def schedule_delayed_delete(path: Path, delay_seconds: int = 900) -> None:
    def worker() -> None:
        time.sleep(delay_seconds)
        secure_delete_tree(path)

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()


def render_legacy_page(status: str = "") -> str:
    escaped_status = html.escape(status)
    public_log_url = html.escape(
        CONFIG.get(
            "public_log_url",
            "/xml/excel_image_server.log",
        ),
        quote=True,
    )
    return f"""<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Image Server</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 0; background: #f6f7f9; color: #17202a; }}
    main {{ max-width: 920px; margin: 0 auto; padding: 28px; }}
    section {{ background: #fff; border: 1px solid #d8dee4; border-radius: 8px; padding: 22px; margin-bottom: 18px; }}
    h1 {{ font-size: 28px; margin: 0 0 18px; }}
    h2 {{ font-size: 18px; margin: 0 0 14px; }}
    label {{ display: block; font-weight: 600; margin-top: 14px; }}
    input, select {{ width: 100%; box-sizing: border-box; padding: 10px; margin-top: 6px; border: 1px solid #b8c0cc; border-radius: 6px; font-size: 15px; }}
    .row {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; }}
    button {{ margin-top: 18px; padding: 12px 16px; border: 0; border-radius: 6px; background: #1167b1; color: white; font-weight: 700; cursor: pointer; }}
    button:hover {{ background: #0d5796; }}
    .status {{ white-space: pre-wrap; background: #eef6ff; border: 1px solid #b7d8f5; border-radius: 6px; padding: 12px; }}
    .muted {{ color: #5d6b7a; }}
  </style>
</head>
<body>
  <main>
    <h1>Excel Image Server</h1>
    <section>
      <h2>Обработка Excel</h2>
      <form action="/process" method="post" enctype="multipart/form-data">
        <label>Excel файл (.xlsx или .xlsm)
          <input name="file" type="file" accept=".xlsx,.xlsm" required>
        </label>
        <label>Лист
          <input name="sheet_name" value="" placeholder="Оставьте пустым, чтобы взять первый лист">
        </label>
        <div class="row">
          <label>Колонка артикула
            <input name="article_column" value="{html.escape(CONFIG.get('default_article_column', 'A'))}" required>
          </label>
          <label>Колонка картинки
            <input name="image_column" value="{html.escape(CONFIG.get('default_image_column', 'B'))}" required>
          </label>
          <label>Начальная строка
            <input name="start_row" type="number" min="1" value="{CONFIG.get('default_start_row', 2)}" required>
          </label>
        </div>
        <button type="submit">Обработать и скачать</button>
      </form>
      <p class="muted">Если лист не указан, сервер возьмет первый лист книги. Для .xlsm макросы сохраняются через keep_vba.</p>
    </section>
    <section>
      <h2>Состояние XML</h2>
      <p><a href="/status">Открыть статус</a> | <a href="{public_log_url}" target="_blank" rel="noopener">Открыть публичный лог</a> | <a href="/reload-index">Перезагрузить XML-индекс</a></p>
      {f'<div class="status">{escaped_status}</div>' if escaped_status else ''}
    </section>
  </main>
</body>
</html>"""


def render_page(status: str = "") -> str:
    default_color = sanitize_hex_color(
        CONFIG.get("default_cell_background_color", "D9D9D9"),
        "D9D9D9",
    )
    default_desired_mb = CONFIG.get("default_desired_output_mb", 20)
    hard_max_output_mb = CONFIG.get("hard_max_output_mb", 100)
    max_local_images_upload_mb = int(
        CONFIG.get("max_local_images_upload_mb", 500)
    )
    return f"""<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Image Server</title>
  <style>
    :root {{ color-scheme: light; }}
    body {{
      font-family: Segoe UI, Arial, sans-serif;
      margin: 0;
      background: #edf2f7;
      color: #111827;
      font-size: 17px;
    }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 34px; }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 22px;
    }}
    h1 {{ font-size: 42px; line-height: 1.05; margin: 0; letter-spacing: 0; }}
    h2 {{ font-size: 24px; margin: 0 0 18px; }}
    section {{
      background: #ffffff;
      border: 1px solid #cbd5e1;
      border-radius: 8px;
      padding: 26px;
      margin-bottom: 20px;
      box-shadow: 0 14px 35px rgba(15, 23, 42, 0.08);
    }}
    label {{ display: block; font-weight: 800; margin-top: 16px; color: #1f2937; }}
    input, select {{
      width: 100%;
      box-sizing: border-box;
      padding: 14px 15px;
      margin-top: 8px;
      border: 2px solid #cbd5e1;
      border-radius: 8px;
      font-size: 17px;
      background: #f8fafc;
      color: #111827;
    }}
    input:focus, select:focus {{
      outline: none;
      border-color: #2563eb;
      background: #ffffff;
      box-shadow: 0 0 0 4px rgba(37, 99, 235, 0.14);
    }}
    input[type="color"] {{ height: 54px; padding: 5px; }}
    input[type="file"] {{
      min-height: 78px;
      padding: 20px;
      border-style: dashed;
      border-width: 3px;
      background: #f0f9ff;
      cursor: pointer;
      font-weight: 800;
    }}
    .file-drop-zone {{
      position: relative;
      display: grid;
      place-items: center;
      min-height: 190px;
      margin-top: 10px;
      padding: 28px;
      overflow: hidden;
      border: 3px dashed #2563eb;
      border-radius: 8px;
      background: #eff6ff;
      color: #0f172a;
      text-align: center;
      box-shadow: inset 0 0 0 1px rgba(37, 99, 235, 0.12);
      transition: background 120ms ease, border-color 120ms ease, box-shadow 120ms ease, transform 120ms ease;
    }}
    .file-drop-zone:hover,
    .file-drop-zone.is-dragover {{
      background: #dbeafe;
      border-color: #0f766e;
      box-shadow: inset 0 0 0 1px rgba(15, 118, 110, 0.24), 0 10px 26px rgba(37, 99, 235, 0.12);
    }}
    .file-drop-zone.is-attached {{
      background: #ecfdf5;
      border-color: #16a34a;
    }}
    .file-drop-zone input[type="file"] {{
      position: absolute;
      inset: 0;
      width: 100%;
      height: 100%;
      min-height: 100%;
      margin: 0;
      padding: 0;
      opacity: 0;
      cursor: pointer;
      z-index: 1;
    }}
    .file-drop-content {{
      pointer-events: none;
      position: relative;
      z-index: 2;
      display: grid;
      gap: 8px;
      justify-items: center;
    }}
    .file-drop-title {{
      font-size: 26px;
      line-height: 1.15;
      font-weight: 950;
      color: #1d4ed8;
    }}
    .file-drop-subtitle {{
      color: #475569;
      font-size: 16px;
      font-weight: 800;
    }}
    .file-selected {{
      margin-top: 10px;
      padding: 10px 14px;
      border-radius: 8px;
      background: #ffffff;
      border: 1px solid #bfdbfe;
      color: #475569;
      font-weight: 900;
      max-width: 100%;
      overflow-wrap: anywhere;
    }}
    .file-selected.is-attached {{
      background: #dcfce7;
      border-color: #86efac;
      color: #166534;
    }}
    .file-clear-button {{
      display: none;
      position: relative;
      z-index: 3;
      width: auto;
      margin-top: 16px;
      padding: 10px 15px;
      border: 1px solid #fecaca;
      border-radius: 8px;
      background: #ffffff;
      color: #b91c1c;
      font-size: 15px;
      font-weight: 900;
      box-shadow: none;
      pointer-events: auto;
    }}
    .file-clear-button.is-visible {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
    }}
    .file-clear-button:hover {{
      background: #fee2e2;
      border-color: #fca5a5;
      box-shadow: none;
      filter: none;
    }}
    .file-clear-button:active {{
      transform: translateY(1px);
      box-shadow: none;
    }}
    .file-clear-button:focus-visible {{
      outline: none;
      box-shadow:
        0 0 0 3px rgba(255, 255, 255, 0.95),
        0 0 0 6px rgba(239, 68, 68, 0.28);
    }}
    .folder-clear-button {{
      display: none;
      width: 100%;
      margin-top: 8px;
      padding: 10px 13px;
      border: 1px solid #fecaca;
      border-radius: 8px;
      background: #ffffff;
      color: #b91c1c;
      font-size: 14px;
      font-weight: 900;
      cursor: pointer;
    }}
    .folder-clear-button.is-visible {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
    }}
    .folder-clear-button:hover {{
      background: #fee2e2;
      border-color: #fca5a5;
    }}
    .field-hint {{
      display: block;
      margin-top: 6px;
      color: #64748b;
      font-size: 14px;
      font-weight: 700;
      line-height: 1.35;
    }}
    .field-hint.is-warning {{
      color: #b45309;
    }}
    .row {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; }}
    .row.two {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    fieldset {{ margin: 22px 0 4px; padding: 0; border: 0; }}
    legend {{ font-size: 20px; font-weight: 900; color: #111827; }}
    .source-list {{ margin-top: 10px; border-top: 1px solid #cbd5e1; }}
    .source-row {{
      display: grid;
      grid-template-columns: minmax(150px, 0.7fr) 120px minmax(260px, 1.8fr);
      gap: 14px;
      align-items: center;
      padding: 14px 0;
      border-bottom: 1px solid #cbd5e1;
    }}
    .source-name {{ font-weight: 900; color: #111827; }}
    .source-row label {{ margin: 0; }}
    .source-row input, .source-row select {{ margin-top: 5px; }}
    .source-detail {{ color: #64748b; font-size: 14px; font-weight: 700; }}
    .folder-picker-label {{ display: block; margin: 0; }}
    .folder-picker-label.is-attached .folder-picker-button {{
      background: #ecfdf5;
      border-color: #16a34a;
      box-shadow: inset 0 0 0 1px rgba(22, 163, 74, 0.12);
    }}
    .folder-picker-label.is-dragover .folder-picker-button {{
      background: #ccfbf1;
      border-color: #2563eb;
      box-shadow: 0 0 0 4px rgba(37, 99, 235, 0.14);
    }}
    .folder-input {{
      position: absolute;
      width: 1px;
      height: 1px;
      min-height: 0;
      margin: 0;
      padding: 0;
      opacity: 0;
      overflow: hidden;
      pointer-events: none;
    }}
    .folder-picker-button {{
      display: grid;
      align-items: center;
      justify-content: center;
      width: 100%;
      min-height: 88px;
      box-sizing: border-box;
      margin-top: 8px;
      padding: 15px;
      border: 3px dashed #2563eb;
      border-radius: 8px;
      background: #eff6ff;
      color: #1d4ed8;
      font-size: 17px;
      font-weight: 900;
      text-align: center;
      cursor: pointer;
      transition: background 120ms ease, border-color 120ms ease, box-shadow 120ms ease, transform 120ms ease;
    }}
    .folder-picker-button:hover {{
      background: #dbeafe;
      border-color: #0f766e;
    }}
    .folder-picker-label:active .folder-picker-button {{
      transform: translateY(1px);
    }}
    .folder-picker-title {{
      display: block;
      line-height: 1.2;
    }}
    .folder-picker-name {{
      display: block;
      max-width: 100%;
      margin-top: 7px;
      color: #475569;
      font-size: 14px;
      line-height: 1.3;
      font-weight: 900;
      overflow-wrap: anywhere;
    }}
    .folder-picker-label.is-attached .folder-picker-title {{
      color: #166534;
    }}
    .folder-picker-label.is-attached .folder-picker-name {{
      color: #14532d;
    }}
    .check-option {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      justify-self: start;
      align-self: end;
      width: fit-content;
      max-width: 100%;
      min-height: 0;
      height: 54px;
      box-sizing: border-box;
      margin-top: 16px;
      padding: 12px 14px;
      border: 2px solid #cbd5e1;
      border-radius: 8px;
      background: #f8fafc;
      cursor: pointer;
    }}
    .check-option input[type="checkbox"] {{
      width: 20px;
      height: 20px;
      min-height: 20px;
      margin: 0;
      padding: 0;
      cursor: pointer;
    }}
    .check-option span {{
      font-weight: 900;
      color: #111827;
      line-height: 1.25;
    }}
    button {{
      width: 100%;
      margin-top: 24px;
      padding: 17px 20px;
      border: 0;
      border-radius: 8px;
      background: #f97316;
      color: white;
      font-size: 20px;
      font-weight: 900;
      cursor: pointer;
      box-shadow: 0 12px 28px rgba(249, 115, 22, 0.25);
      transform: translateY(0);
      transition:
        transform 120ms ease,
        box-shadow 120ms ease,
        background 120ms ease,
        filter 120ms ease;
    }}
    button:hover {{
      background: #ea580c;
      box-shadow: 0 15px 32px rgba(249, 115, 22, 0.34);
      filter: brightness(1.02);
    }}
    button:active {{
      transform: translateY(2px) scale(0.995);
      box-shadow: 0 6px 16px rgba(249, 115, 22, 0.24);
      filter: brightness(0.96);
    }}
    button:focus-visible {{
      outline: none;
      box-shadow:
        0 0 0 4px rgba(255, 255, 255, 0.95),
        0 0 0 8px rgba(249, 115, 22, 0.4),
        0 12px 28px rgba(249, 115, 22, 0.25);
    }}
    button:disabled {{
      background: #94a3b8;
      cursor: progress;
      box-shadow: none;
      transform: none;
      filter: none;
    }}
    .cancel-button {{
      background: #dc2626;
      box-shadow: 0 12px 28px rgba(220, 38, 38, 0.22);
    }}
    .cancel-button:hover {{
      background: #b91c1c;
      box-shadow: 0 15px 32px rgba(220, 38, 38, 0.3);
    }}
    .cancel-button:disabled {{
      background: #94a3b8;
      cursor: progress;
      box-shadow: none;
    }}
    a {{ color: #1d4ed8; font-weight: 800; }}
    .pill {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 10px 14px;
      background: #dcfce7;
      border: 1px solid #86efac;
      border-radius: 8px;
      font-weight: 800;
      color: #166534;
    }}
    .status {{ white-space: pre-wrap; background: #eff6ff; border: 1px solid #93c5fd; border-radius: 8px; padding: 14px; }}
    .muted {{ color: #64748b; font-weight: 600; }}
    .sheet-preview {{ margin-top: 18px; }}
    .sheet-preview h2 {{ margin-top: 0; }}
    .progress-panel {{
      display: none;
      margin-top: 16px;
      padding: 14px;
      border: 1px solid #bfdbfe;
      border-radius: 8px;
      background: #eff6ff;
    }}
    .progress-panel.is-visible {{ display: block; }}
    .progress-top {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 10px;
      color: #1e3a8a;
      font-weight: 900;
    }}
    .progress-track {{
      width: 100%;
      height: 16px;
      overflow: hidden;
      border-radius: 999px;
      background: #dbeafe;
      box-shadow: inset 0 0 0 1px rgba(37, 99, 235, 0.16);
    }}
    .progress-fill {{
      width: 0%;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #2563eb, #0f766e);
      transition: width 180ms ease;
    }}
    .processing-report {{
      display: none;
      margin-top: 18px;
      border: 1px solid #cbd5e1;
      border-radius: 8px;
      background: #ffffff;
      overflow: hidden;
    }}
    .processing-report.is-visible {{ display: block; }}
    .processing-report summary {{
      padding: 14px 16px;
      cursor: pointer;
      background: #f8fafc;
      color: #0f172a;
      font-size: 18px;
      font-weight: 900;
    }}
    .report-content {{
      display: grid;
      gap: 16px;
      padding: 16px;
    }}
    .report-stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}
    .report-chip {{
      padding: 9px 12px;
      border-radius: 8px;
      background: #eff6ff;
      color: #1e3a8a;
      font-weight: 900;
    }}
    .report-content h3 {{
      margin: 0 0 8px;
      color: #111827;
      font-size: 16px;
    }}
    .report-content ol,
    .report-content ul {{
      margin: 0;
      padding-left: 24px;
    }}
    .report-missing-list {{
      max-height: 260px;
      overflow: auto;
      border: 1px solid #e2e8f0;
      border-radius: 8px;
      padding: 10px 10px 10px 28px;
      background: #f8fafc;
    }}
    .report-empty {{ margin: 0; color: #64748b; font-weight: 700; }}
    .preview-wrap {{ overflow: auto; border: 1px solid #cbd5e1; border-radius: 8px; background: #ffffff; max-height: 420px; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 620px; }}
    td, th {{ border: 1px solid #e2e8f0; padding: 9px 11px; max-width: 260px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
    th {{ background: #0f766e; color: white; position: sticky; top: 0; z-index: 1; }}
    tr:nth-child(even) td {{ background: #f8fafc; }}
    .preview-meta {{ margin: 0 0 12px; color: #475569; font-weight: 800; }}
    @media (max-width: 760px) {{
      main {{ padding: 18px; }}
      h1 {{ font-size: 31px; }}
      .topbar, .row, .row.two {{ display: block; }}
      .source-row {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <main>
    <div class="topbar">
      <h1>Excel Image Server</h1>
    </div>
    <section>
      <h2>Обробка Excel</h2>
      <form id="processForm" action="/process" method="post" enctype="multipart/form-data">
        <label>Excel-файл (.xlsx або .xlsm)
          <div id="fileDropZone" class="file-drop-zone">
            <input id="fileInput" name="file" type="file" accept=".xlsx,.xlsm" required>
            <div class="file-drop-content">
              <div class="file-drop-title">Перетягніть Excel-файл сюди</div>
              <div class="file-drop-subtitle">або натисніть у цю область, щоб вибрати файл</div>
              <div id="fileSelectedInfo" class="file-selected">Файл ще не прикріплено</div>
            </div>
            <button id="clearFileButton" class="file-clear-button" type="button">Відкріпити файл</button>
          </div>
          <span class="field-hint">Якщо файл посилається на зовнішні джерела через ВПР, XLOOKUP або подібні формули, Excel може відкривати результат через відновлення; зовнішні посилання можуть втратитися, але збережені значення мають лишитися.</span>
        </label>
        <label>Аркуш
          <select id="sheetSelect" name="sheet_name">
            <option value="">Спочатку виберіть Excel-файл</option>
          </select>
        </label>
        <div class="sheet-preview">
          <h2>Попередній перегляд даних</h2>
          <p id="previewMeta" class="preview-meta">Файл ще не вибрано.</p>
          <div class="preview-wrap" id="previewWrap"></div>
        </div>
        <fieldset>
          <legend>Фото</legend>
          <div class="source-list">
            <div class="source-row">
              <div>
                <div class="source-name">Основні фото</div>
                <div class="source-detail">Фото з каталогу</div>
              </div>
              <label>Пріоритет
                <select class="priority-select" data-source="server" data-previous="1">
                  <option value="1" selected>1</option>
                  <option value="2">2</option>
                  <option value="3">3</option>
                  <option value="off">Вимкнено</option>
                </select>
              </label>
              <div class="source-detail">Доступна завжди</div>
            </div>
            <div class="source-row">
              <div>
                <div class="source-name">Вибрана папка 1</div>
                <div class="source-detail" id="local1Count">Папку не вибрано</div>
              </div>
              <label>Пріоритет
                <select class="priority-select" data-source="local_1" data-previous="2">
                  <option value="1">1</option>
                  <option value="2" selected>2</option>
                  <option value="3">3</option>
                  <option value="off">Вимкнено</option>
                </select>
              </label>
              <label class="folder-picker-label">Папка
                <input id="local1Input" class="folder-input" type="file" accept="image/*" webkitdirectory directory multiple>
                <span class="folder-picker-button">
                  <span class="folder-picker-title">Вибрати папку або перетягнути сюди</span>
                  <span class="folder-picker-name">Папку не прикріплено</span>
                </span>
                <button id="clearLocal1Button" class="folder-clear-button" type="button">Відкріпити папку</button>
              </label>
            </div>
            <div class="source-row">
              <div>
                <div class="source-name">Вибрана папка 2</div>
                <div class="source-detail" id="local2Count">Папку не вибрано</div>
              </div>
              <label>Пріоритет
                <select class="priority-select" data-source="local_2" data-previous="3">
                  <option value="1">1</option>
                  <option value="2">2</option>
                  <option value="3" selected>3</option>
                  <option value="off">Вимкнено</option>
                </select>
              </label>
              <label class="folder-picker-label">Папка
                <input id="local2Input" class="folder-input" type="file" accept="image/*" webkitdirectory directory multiple>
                <span class="folder-picker-button">
                  <span class="folder-picker-title">Вибрати папку або перетягнути сюди</span>
                  <span class="folder-picker-name">Папку не прикріплено</span>
                </span>
                <button id="clearLocal2Button" class="folder-clear-button" type="button">Відкріпити папку</button>
              </label>
            </div>
          </div>
        </fieldset>
        <div class="row">
          <label>Стовпець артикула
            <input name="article_column" value="{html.escape(CONFIG.get('default_article_column', 'A'))}" placeholder="A або 1" required>
            <span class="field-hint">Можна ввести літеру Excel-стовпця або номер з 1.</span>
          </label>
          <label>Стовпець картинки
            <input name="image_column" value="{html.escape(CONFIG.get('default_image_column', 'B'))}" placeholder="B або 2" required>
            <span class="field-hint">Картинка буде вставлена в цей стовпець.</span>
          </label>
          <label>Початковий рядок
            <input name="start_row" type="number" min="1" value="{CONFIG.get('default_start_row', 2)}" required>
          </label>
        </div>
        <div class="row">
          <label>Фон комірки з картинкою
            <input name="cell_background_color" type="color" value="#{default_color}">
          </label>
          <label class="check-option">
            <input name="skip_cell_background" type="checkbox" value="true">
            <span>Не змінювати фон комірки</span>
          </label>
          <label>Бажаний розмір результату, МБ
            <input name="desired_output_mb" type="number" min="1" max="100" step="1" value="{default_desired_mb}" required>
          </label>
        </div>
        <button id="submitButton" type="submit">Обробити файл</button>
        <button id="cancelButton" class="cancel-button" type="button" disabled>Скасувати обробку</button>
        <p id="processStatus" class="muted" aria-live="polite"></p>
        <div id="progressPanel" class="progress-panel" aria-live="polite">
          <div class="progress-top">
            <span id="progressMessage">Очікування...</span>
            <span id="progressPercent">0%</span>
          </div>
          <div class="progress-track">
            <div id="progressFill" class="progress-fill"></div>
          </div>
        </div>
        <details id="processingReport" class="processing-report">
          <summary>Детальна інформація</summary>
          <div id="reportContent" class="report-content"></div>
        </details>
      </form>
    </section>
  </main>
  <script>
    window.EXCEL_IMAGE_SERVER_CONFIG = {{
      maxUploadMb: {CONFIG['max_upload_mb']},
      maxLocalImagesUploadMb: {max_local_images_upload_mb},
      defaultSheetMessage: 'Спочатку виберіть Excel-файл'
    }};
  </script>
  <script src="/client.js"></script>
</body>
</html>"""


CLIENT_JS = r"""
const fileInput = document.getElementById('fileInput');
const fileDropZone = document.getElementById('fileDropZone');
const fileSelectedInfo = document.getElementById('fileSelectedInfo');
const clearFileButton = document.getElementById('clearFileButton');
const processForm = document.getElementById('processForm');
const submitButton = document.getElementById('submitButton');
const cancelButton = document.getElementById('cancelButton');
const sheetSelect = document.getElementById('sheetSelect');
const previewWrap = document.getElementById('previewWrap');
const previewMeta = document.getElementById('previewMeta');
const processStatus = document.getElementById('processStatus');
const progressPanel = document.getElementById('progressPanel');
const progressMessage = document.getElementById('progressMessage');
const progressPercent = document.getElementById('progressPercent');
const progressFill = document.getElementById('progressFill');
const processingReport = document.getElementById('processingReport');
const reportContent = document.getElementById('reportContent');
const backgroundColorInput = processForm.elements.cell_background_color;
const skipBackgroundInput = processForm.elements.skip_cell_background;
const desiredOutputInput = processForm.elements.desired_output_mb;
const localInputs = {
  local_1: document.getElementById('local1Input'),
  local_2: document.getElementById('local2Input'),
};
const localClearButtons = {
  local_1: document.getElementById('clearLocal1Button'),
  local_2: document.getElementById('clearLocal2Button'),
};
const localDropLabels = {
  local_1: document.getElementById('local1Input').closest('.folder-picker-label'),
  local_2: document.getElementById('local2Input').closest('.folder-picker-label'),
};
const localCounts = {
  local_1: document.getElementById('local1Count'),
  local_2: document.getElementById('local2Count'),
};
const prioritySelects = Array.from(
  document.querySelectorAll('.priority-select')
);
const supportedImageExtensions = new Set([
  '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp'
]);
const maxLocalUploadBytes =
  (window.EXCEL_IMAGE_SERVER_CONFIG?.maxLocalImagesUploadMb || 500) * 1024 * 1024;
const maxExcelUploadBytes =
  (window.EXCEL_IMAGE_SERVER_CONFIG?.maxUploadMb || 100) * 1024 * 1024;
const settingsKey = 'excel-image-server-settings-v3';
const utf8Decoder = new TextDecoder('utf-8');
let workbookState = null;
let progressTimer = null;
let heartbeatTimer = null;
let activeProgressId = '';
let cancelRequested = false;
const localFolderState = {
  local_1: { files: [], folderName: '' },
  local_2: { files: [], folderName: '' },
};
const localRelativePaths = new WeakMap();

function saveSettings() {
  const priorities = {};
  prioritySelects.forEach((select) => {
    priorities[select.dataset.source] = select.value;
  });
  const settings = {
    sheet_name: sheetSelect.value || '',
    article_column: processForm.elements.article_column.value,
    image_column: processForm.elements.image_column.value,
    start_row: processForm.elements.start_row.value,
    cell_background_color: backgroundColorInput.value,
    skip_cell_background: skipBackgroundInput.checked,
    desired_output_mb: processForm.elements.desired_output_mb.value,
    priorities,
  };
  localStorage.setItem(settingsKey, JSON.stringify(settings));
}

function loadSettings() {
  try {
    return JSON.parse(localStorage.getItem(settingsKey) || '{}');
  } catch (_) {
    return {};
  }
}

function applySavedSettings() {
  const settings = loadSettings();
  const fieldNames = [
    'article_column',
    'image_column',
    'start_row',
    'cell_background_color',
    'desired_output_mb',
  ];
  fieldNames.forEach((name) => {
    if (settings[name] !== undefined && processForm.elements[name]) {
      processForm.elements[name].value = settings[name];
    }
  });
  if (settings.skip_cell_background !== undefined) {
    skipBackgroundInput.checked = Boolean(settings.skip_cell_background);
  }
  if (settings.priorities) {
    prioritySelects.forEach((select) => {
      const savedValue = settings.priorities[select.dataset.source];
      if (['1', '2', '3', 'off'].includes(savedValue)) {
        select.value = savedValue;
        select.dataset.previous = savedValue;
      }
    });
  }
}

applySavedSettings();

function updateBackgroundControl() {
  backgroundColorInput.disabled = skipBackgroundInput.checked;
}

updateBackgroundControl();

function formatFileSize(bytes) {
  if (!Number.isFinite(bytes)) return '';
  const megabytes = bytes / (1024 * 1024);
  if (megabytes >= 1) return megabytes.toFixed(1) + ' МБ';
  return Math.max(1, Math.round(bytes / 1024)) + ' КБ';
}

function ensureExcelFileSizeAllowed(file) {
  if (file && file.size > maxExcelUploadBytes) {
    throw new Error(
      'Excel-файл більший за ліміт ' +
      (window.EXCEL_IMAGE_SERVER_CONFIG?.maxUploadMb || 100) + ' МБ.'
    );
  }
}

async function fileToBase64(file) {
  const bytes = new Uint8Array(await file.arrayBuffer());
  const chunkSize = 0x8000;
  let binary = '';
  for (let index = 0; index < bytes.length; index += chunkSize) {
    binary += String.fromCharCode.apply(
      null,
      bytes.subarray(index, index + chunkSize)
    );
  }
  return btoa(binary);
}

function clampDesiredSize(value) {
  return Math.max(1, Math.min(100, Math.ceil(Number(value) || 1)));
}

function updateAttachedFileState(file) {
  if (file) {
    fileDropZone.classList.add('is-attached');
    fileSelectedInfo.classList.add('is-attached');
    clearFileButton.classList.add('is-visible');
    fileSelectedInfo.textContent =
      'Прикріплено: ' + file.name + ' (' + formatFileSize(file.size) + ')';
  } else {
    fileDropZone.classList.remove('is-attached');
    fileSelectedInfo.classList.remove('is-attached');
    clearFileButton.classList.remove('is-visible');
    fileSelectedInfo.textContent = 'Файл ще не прикріплено';
  }
}

function resetAttachedFile() {
  fileInput.value = '';
  workbookState = null;
  sheetSelect.innerHTML = '<option value="">Спочатку виберіть Excel-файл</option>';
  previewMeta.textContent = 'Файл ще не вибрано.';
  previewWrap.innerHTML = '';
  processStatus.textContent = '';
  stopProgressPolling();
  progressPanel.classList.remove('is-visible');
  progressFill.style.width = '0%';
  progressPercent.textContent = '0%';
  progressMessage.textContent = 'Очікування...';
  hideProcessingReport();
  updateAttachedFileState(null);
}

function hideProcessingReport() {
  processingReport.classList.remove('is-visible');
  processingReport.open = false;
  reportContent.replaceChildren();
}

function isExcelFile(file) {
  return /\.(xlsx|xlsm)$/i.test(file?.name || '');
}

['dragenter', 'dragover'].forEach((eventName) => {
  fileDropZone.addEventListener(eventName, (event) => {
    event.preventDefault();
    fileDropZone.classList.add('is-dragover');
  });
});

['dragleave', 'drop'].forEach((eventName) => {
  fileDropZone.addEventListener(eventName, () => {
    fileDropZone.classList.remove('is-dragover');
  });
});

fileDropZone.addEventListener('drop', (event) => {
  event.preventDefault();
  const files = event.dataTransfer?.files;
  if (!files || !files.length) return;
  if (!isExcelFile(files[0])) {
    processStatus.textContent = 'Можна прикріпити тільки .xlsx або .xlsm файл.';
    return;
  }
  try {
    ensureExcelFileSizeAllowed(files[0]);
    fileInput.files = files;
    fileInput.dispatchEvent(new Event('change', { bubbles: true }));
  } catch (error) {
    processStatus.textContent = error.message || 'Не вдалося прикріпити файл перетягуванням. Натисніть на область і виберіть файл.';
  }
});

clearFileButton.addEventListener('click', (event) => {
  event.preventDefault();
  event.stopPropagation();
  resetAttachedFile();
});

prioritySelects.forEach((select) => {
  select.addEventListener('change', () => {
    const previous = select.dataset.previous;
    const selected = select.value;
    const other = selected === 'off' ? null : prioritySelects.find(
      (candidate) => candidate !== select && candidate.value === selected
    );
    if (other && previous) {
      other.value = previous;
      other.dataset.previous = previous;
    }
    select.dataset.previous = selected;
    saveSettings();
  });
});

[
  processForm.elements.article_column,
  processForm.elements.image_column,
  processForm.elements.start_row,
  backgroundColorInput,
  skipBackgroundInput,
].forEach((input) => input.addEventListener('change', () => {
  updateBackgroundControl();
  saveSettings();
}));
desiredOutputInput.addEventListener('change', () => {
  desiredOutputInput.value = clampDesiredSize(desiredOutputInput.value);
  saveSettings();
});
sheetSelect.addEventListener('change', () => {
  saveSettings();
  renderSelectedPreview();
});

Object.entries(localInputs).forEach(([source, input]) => {
  input.addEventListener('change', () => {
    updateLocalFolderState(source, Array.from(input.files || []), '');
  });
});

Object.entries(localClearButtons).forEach(([source, button]) => {
  button.addEventListener('click', (event) => {
    event.preventDefault();
    event.stopPropagation();
    clearLocalFolder(source);
  });
});

function getLocalFiles(source) {
  const draggedFiles = localFolderState[source]?.files || [];
  return draggedFiles.length ? draggedFiles : Array.from(localInputs[source].files || []);
}

function updateLocalFolderState(source, files, folderName) {
  const imageFiles = files.filter((file) =>
    supportedImageExtensions.has(fileExtension(file.name))
  );
  const resolvedFolderName = folderName || inferFolderName(imageFiles);
  localFolderState[source] = {
    files: imageFiles,
    folderName: resolvedFolderName,
  };
  const label = localDropLabels[source];
  const title = label?.querySelector('.folder-picker-title');
  const name = label?.querySelector('.folder-picker-name');
  const clearButton = localClearButtons[source];
  label?.classList.toggle('is-attached', imageFiles.length > 0);
  clearButton?.classList.toggle('is-visible', imageFiles.length > 0);
  if (title) {
    title.textContent = imageFiles.length
      ? 'Папку прикріплено'
      : 'Вибрати папку або перетягнути сюди';
  }
  if (name) {
    name.textContent = imageFiles.length
      ? (resolvedFolderName || 'Вибрана папка') + ' · зображень: ' + imageFiles.length
      : 'Папку не прикріплено';
  }
  localCounts[source].textContent = imageFiles.length
    ? 'Папка: ' + (resolvedFolderName || 'Вибрана папка') + ' · зображень: ' + imageFiles.length
    : 'Папку не вибрано';
}

function clearLocalFolder(source) {
  localInputs[source].value = '';
  localFolderState[source] = { files: [], folderName: '' };
  updateLocalFolderState(source, [], '');
  processStatus.textContent = 'Папку відкріплено.';
}

function inferFolderName(files) {
  const firstFile = files[0];
  const relativePath =
    firstFile?.webkitRelativePath || localRelativePaths.get(firstFile) || '';
  if (relativePath.includes('/')) {
    return relativePath.split('/')[0] || '';
  }
  return '';
}

function readDirectoryEntries(reader) {
  return new Promise((resolve, reject) => {
    reader.readEntries(resolve, reject);
  });
}

async function collectEntryFiles(entry, prefix = '') {
  if (!entry) return [];
  if (entry.isFile) {
    return await new Promise((resolve) => {
      entry.file((file) => {
        localRelativePaths.set(file, prefix + file.name);
        resolve([file]);
      }, () => resolve([]));
    });
  }
  if (!entry.isDirectory) return [];
  const reader = entry.createReader();
  const files = [];
  while (true) {
    const entries = await readDirectoryEntries(reader);
    if (!entries.length) break;
    for (const child of entries) {
      files.push(...await collectEntryFiles(child, prefix + entry.name + '/'));
    }
  }
  return files;
}

async function filesFromDrop(event) {
  const items = Array.from(event.dataTransfer?.items || []);
  if (!items.length) return Array.from(event.dataTransfer?.files || []);
  const files = [];
  for (const item of items) {
    const entry = item.webkitGetAsEntry?.();
    if (entry) {
      files.push(...await collectEntryFiles(entry));
      continue;
    }
    const file = item.getAsFile?.();
    if (file) files.push(file);
  }
  return files;
}

Object.entries(localDropLabels).forEach(([source, label]) => {
  ['dragenter', 'dragover'].forEach((eventName) => {
    label.addEventListener(eventName, (event) => {
      event.preventDefault();
      event.stopPropagation();
      label.classList.add('is-dragover');
    });
  });
  ['dragleave', 'drop'].forEach((eventName) => {
    label.addEventListener(eventName, (event) => {
      event.preventDefault();
      event.stopPropagation();
      label.classList.remove('is-dragover');
    });
  });
  label.addEventListener('drop', async (event) => {
    event.preventDefault();
    event.stopPropagation();
    const files = await filesFromDrop(event);
    updateLocalFolderState(source, files, '');
    const selectedImages = getLocalFiles(source).length;
    processStatus.textContent = selectedImages
      ? 'Папку прикріплено. Знайдено зображень: ' + selectedImages
      : 'У перетягнутій папці не знайдено підтримуваних зображень.';
  });
});

function sourceOrder() {
  return prioritySelects
    .slice()
    .filter((select) => select.value !== 'off')
    .sort((left, right) => Number(left.value) - Number(right.value))
    .map((select) => select.dataset.source);
}

function selectedFolderName(source, fallback) {
  return localFolderState[source]?.folderName || fallback;
}

function searchDirectoryLabels(order) {
  const labels = {
    server: 'Основні фото',
    local_1: selectedFolderName('local_1', 'Вибрана папка 1'),
    local_2: selectedFolderName('local_2', 'Вибрана папка 2'),
  };
  return order.map((source) => labels[source] || source);
}

function appendHeading(parent, text) {
  const heading = document.createElement('h3');
  heading.textContent = text;
  parent.appendChild(heading);
}

function appendList(parent, items, ordered, className) {
  const list = document.createElement(ordered ? 'ol' : 'ul');
  if (className) list.className = className;
  items.forEach((item) => {
    const element = document.createElement('li');
    element.textContent = item;
    list.appendChild(element);
  });
  parent.appendChild(list);
}

async function loadProcessingReport(progressId) {
  try {
    const response = await fetch('/progress/' + encodeURIComponent(progressId), {
      cache: 'no-store',
    });
    const payload = await response.json();
    return payload?.report || null;
  } catch (_) {
    return null;
  }
}

function renderProcessingReport(report, order) {
  if (!report) {
    hideProcessingReport();
    return;
  }
  const notFound = Array.isArray(report.not_found) ? report.not_found : [];
  const failed = Array.isArray(report.failed) ? report.failed : [];
  reportContent.replaceChildren();

  const stats = document.createElement('div');
  stats.className = 'report-stats';
  const inserted = document.createElement('div');
  inserted.className = 'report-chip';
  inserted.textContent = 'Успішно вставлено: ' + (Number(report.inserted) || 0);
  const missing = document.createElement('div');
  missing.className = 'report-chip';
  missing.textContent = 'Не знайдено: ' + (Number(report.not_found_count) || notFound.length);
  stats.append(inserted, missing);
  if (failed.length) {
    const failedChip = document.createElement('div');
    failedChip.className = 'report-chip';
    failedChip.textContent = 'Не вставлено через помилку: ' + failed.length;
    stats.appendChild(failedChip);
  }
  reportContent.appendChild(stats);

  const directories = document.createElement('div');
  appendHeading(directories, 'Директорії пошуку');
  appendList(directories, searchDirectoryLabels(order), true);
  reportContent.appendChild(directories);

  const missingSection = document.createElement('div');
  appendHeading(missingSection, 'Список не знайдених');
  if (notFound.length) {
    appendList(missingSection, notFound, false, 'report-missing-list');
  } else {
    const empty = document.createElement('p');
    empty.className = 'report-empty';
    empty.textContent = 'Немає.';
    missingSection.appendChild(empty);
  }
  reportContent.appendChild(missingSection);

  if (failed.length) {
    const failedSection = document.createElement('div');
    appendHeading(failedSection, 'Не вставлено через помилку');
    appendList(failedSection, failed, false, 'report-missing-list');
    reportContent.appendChild(failedSection);
  }

  processingReport.classList.add('is-visible');
  processingReport.open = true;
}

function normalizeLocalArticle(value, forExcel) {
  let normalized = '';
  for (const character of String(value ?? '').trim()) {
    if (/^[\p{L}\p{N}]$/u.test(character) || character === ' ') {
      normalized += character;
    } else if (!forExcel && character === '_') {
      normalized += character;
    } else {
      normalized += '-';
    }
  }
  return normalized.toLowerCase();
}

function fileStem(filename) {
  const dot = filename.lastIndexOf('.');
  return dot > 0 ? filename.slice(0, dot) : filename;
}

function fileExtension(filename) {
  const dot = filename.lastIndexOf('.');
  return dot >= 0 ? filename.slice(dot).toLowerCase() : '';
}

function buildFolderIndex(source) {
  const index = new Map();
  for (const file of getLocalFiles(source)) {
    if (!supportedImageExtensions.has(fileExtension(file.name))) continue;
    const key = normalizeLocalArticle(fileStem(file.name), false);
    if (key && !index.has(key)) index.set(key, file);
  }
  return index;
}

function readU16(view, offset) {
  return view.getUint16(offset, true);
}

function readU32(view, offset) {
  return view.getUint32(offset, true);
}

function findEndOfCentralDirectory(view) {
  const minimumOffset = Math.max(0, view.byteLength - 66000);
  for (let offset = view.byteLength - 22; offset >= minimumOffset; offset -= 1) {
    if (readU32(view, offset) === 0x06054b50) return offset;
  }
  throw new Error('Не вдалося прочитати структуру Excel-файлу.');
}

async function inflateRaw(bytes) {
  if (!('DecompressionStream' in window)) {
    throw new Error('Браузер не підтримує локальний перегляд .xlsx/.xlsm.');
  }
  const stream = new Blob([bytes]).stream().pipeThrough(
    new DecompressionStream('deflate-raw')
  );
  return new Uint8Array(await new Response(stream).arrayBuffer());
}

async function openZip(buffer) {
  const bytes = new Uint8Array(buffer);
  const view = new DataView(buffer);
  const eocd = findEndOfCentralDirectory(view);
  const entriesCount = readU16(view, eocd + 10);
  let offset = readU32(view, eocd + 16);
  const entries = new Map();

  for (let index = 0; index < entriesCount; index += 1) {
    if (readU32(view, offset) !== 0x02014b50) {
      throw new Error('Пошкоджена центральна таблиця Excel-файлу.');
    }
    const method = readU16(view, offset + 10);
    const compressedSize = readU32(view, offset + 20);
    const nameLength = readU16(view, offset + 28);
    const extraLength = readU16(view, offset + 30);
    const commentLength = readU16(view, offset + 32);
    const localOffset = readU32(view, offset + 42);
    const name = utf8Decoder.decode(bytes.slice(offset + 46, offset + 46 + nameLength));
    entries.set(name, { method, compressedSize, localOffset });
    offset += 46 + nameLength + extraLength + commentLength;
  }

  return {
    async read(name) {
      const entry = entries.get(name);
      if (!entry) throw new Error('У файлі Excel немає частини: ' + name);
      const local = entry.localOffset;
      if (readU32(view, local) !== 0x04034b50) {
        throw new Error('Пошкоджений локальний запис Excel-файлу.');
      }
      const nameLength = readU16(view, local + 26);
      const extraLength = readU16(view, local + 28);
      const start = local + 30 + nameLength + extraLength;
      const compressed = bytes.slice(start, start + entry.compressedSize);
      if (entry.method === 0) return compressed;
      if (entry.method === 8) return await inflateRaw(compressed);
      throw new Error('Непідтримуваний метод стиснення в Excel-файлі.');
    },
    async text(name) {
      return utf8Decoder.decode(await this.read(name));
    },
    has(name) {
      return entries.has(name);
    },
  };
}

function parseXml(text) {
  const xml = new DOMParser().parseFromString(text, 'application/xml');
  if (xml.querySelector('parsererror')) {
    throw new Error('Не вдалося прочитати XML всередині Excel-файлу.');
  }
  return xml;
}

function normalizeZipPath(baseFile, target) {
  const cleanTarget = String(target || '').replaceAll('\\', '/');
  if (cleanTarget.startsWith('/')) return cleanTarget.slice(1);
  const baseDir = baseFile.slice(0, baseFile.lastIndexOf('/') + 1);
  const parts = (baseDir + cleanTarget).split('/');
  const normalized = [];
  for (const part of parts) {
    if (!part || part === '.') continue;
    if (part === '..') normalized.pop();
    else normalized.push(part);
  }
  return normalized.join('/');
}

async function readSharedStrings(zip) {
  if (!zip.has('xl/sharedStrings.xml')) return [];
  const xml = parseXml(await zip.text('xl/sharedStrings.xml'));
  return Array.from(xml.getElementsByTagName('si')).map((item) =>
    Array.from(item.getElementsByTagName('t')).map((node) => node.textContent || '').join('')
  );
}

async function analyzeWorkbook(file) {
  const zip = await openZip(await file.arrayBuffer());
  const workbookXml = parseXml(await zip.text('xl/workbook.xml'));
  const relsXml = parseXml(await zip.text('xl/_rels/workbook.xml.rels'));
  const rels = new Map();
  Array.from(relsXml.getElementsByTagName('Relationship')).forEach((rel) => {
    rels.set(rel.getAttribute('Id'), rel.getAttribute('Target'));
  });
  const sharedStrings = await readSharedStrings(zip);
  const sheets = Array.from(workbookXml.getElementsByTagName('sheet')).map((sheet) => {
    const relationshipId =
      sheet.getAttribute('r:id') ||
      sheet.getAttributeNS('http://schemas.openxmlformats.org/officeDocument/2006/relationships', 'id');
    return {
      name: sheet.getAttribute('name') || '',
      path: normalizeZipPath('xl/workbook.xml', rels.get(relationshipId)),
    };
  }).filter((sheet) => sheet.name && sheet.path);
  return { zip, sharedStrings, sheets, parsedSheets: new Map() };
}

function columnLettersToIndex(letters) {
  let result = 0;
  for (const character of letters.toUpperCase()) {
    result = result * 26 + character.charCodeAt(0) - 64;
  }
  return result;
}

function columnToIndex(value) {
  const text = String(value || '').trim();
  if (!text) throw new Error('Вкажіть стовпець.');
  if (/^\d+$/.test(text)) {
    const number = Number(text);
    if (number < 1) throw new Error('Номер стовпця має починатися з 1.');
    return number;
  }
  if (!/^[A-Za-z]+$/.test(text)) {
    throw new Error('Стовпець можна ввести як літери A, B, C або як номер 1, 2, 3.');
  }
  return columnLettersToIndex(text);
}

function parseCellReference(reference) {
  const match = String(reference || '').match(/^([A-Za-z]+)(\d+)$/);
  if (!match) return null;
  return {
    column: columnLettersToIndex(match[1]),
    row: Number(match[2]),
  };
}

function cellText(cell, sharedStrings) {
  const formula = cell.getElementsByTagName('f')[0];
  if (formula) return '=' + (formula.textContent || '');
  const type = cell.getAttribute('t');
  if (type === 'inlineStr') {
    return Array.from(cell.getElementsByTagName('t')).map((node) => node.textContent || '').join('');
  }
  const valueNode = cell.getElementsByTagName('v')[0];
  const value = valueNode ? valueNode.textContent || '' : '';
  if (type === 's') return sharedStrings[Number(value)] || '';
  if (type === 'b') return value === '1' ? 'TRUE' : 'FALSE';
  return value;
}

function maxFromDimension(xml) {
  const dimension = xml.getElementsByTagName('dimension')[0]?.getAttribute('ref') || '';
  const last = dimension.split(':').pop();
  const parsed = parseCellReference(last);
  return parsed || { row: 0, column: 0 };
}

async function parseSheet(sheetName) {
  const cached = workbookState?.parsedSheets.get(sheetName);
  if (cached) return cached;
  const sheet = workbookState?.sheets.find((item) => item.name === sheetName);
  if (!sheet) throw new Error('Аркуш не знайдено.');
  const xml = parseXml(await workbookState.zip.text(sheet.path));
  const dimension = maxFromDimension(xml);
  const rows = new Map();
  let maxRow = dimension.row;
  let maxColumn = dimension.column;
  Array.from(xml.getElementsByTagName('c')).forEach((cell) => {
    const reference = parseCellReference(cell.getAttribute('r'));
    if (!reference) return;
    maxRow = Math.max(maxRow, reference.row);
    maxColumn = Math.max(maxColumn, reference.column);
    const value = cellText(cell, workbookState.sharedStrings);
    if (!rows.has(reference.row)) rows.set(reference.row, new Map());
    rows.get(reference.row).set(reference.column, value);
  });
  const parsed = { rows, maxRow, maxColumn };
  workbookState.parsedSheets.set(sheetName, parsed);
  return parsed;
}

function escapeHtml(value) {
  return String(value)
    .replaceAll('&', '&amp;')
    .replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;')
    .replaceAll('"', '&quot;')
    .replaceAll("'", '&#039;');
}

async function renderSelectedPreview() {
  const sheetName = sheetSelect.value;
  if (!workbookState || !sheetName) {
    previewMeta.textContent = 'Файл ще не вибрано.';
    previewWrap.innerHTML = '';
    return;
  }
  try {
    const parsed = await parseSheet(sheetName);
    const previewRows = Math.min(parsed.maxRow || 1, 12);
    const previewCols = Math.min(parsed.maxColumn || 1, 12);
    previewMeta.textContent =
      sheetName + ': рядків ' + parsed.maxRow + ', стовпців ' + parsed.maxColumn;
    let htmlTable = '<table><thead><tr><th>№</th>';
    for (let column = 1; column <= previewCols; column += 1) {
      htmlTable += '<th>' + column + '</th>';
    }
    htmlTable += '</tr></thead><tbody>';
    for (let row = 1; row <= previewRows; row += 1) {
      htmlTable += '<tr><th>' + row + '</th>';
      const rowValues = parsed.rows.get(row) || new Map();
      for (let column = 1; column <= previewCols; column += 1) {
        htmlTable += '<td>' + escapeHtml(rowValues.get(column) || '') + '</td>';
      }
      htmlTable += '</tr>';
    }
    htmlTable += '</tbody></table>';
    previewWrap.innerHTML = htmlTable;
  } catch (error) {
    previewMeta.textContent = error.message;
    previewWrap.innerHTML = '';
  }
}

fileInput.addEventListener('change', async () => {
  const file = fileInput.files[0];
  updateAttachedFileState(file);
  workbookState = null;
  hideProcessingReport();
  sheetSelect.innerHTML = '<option value="">Завантаження аркушів...</option>';
  previewMeta.textContent = 'Читаю Excel-файл...';
  previewWrap.innerHTML = '';
  processStatus.textContent = '';
  if (!file) {
    updateAttachedFileState(null);
    sheetSelect.innerHTML = '<option value="">Спочатку виберіть Excel-файл</option>';
    previewMeta.textContent = 'Файл ще не вибрано.';
    return;
  }
  if (!isExcelFile(file)) {
    fileInput.value = '';
    updateAttachedFileState(null);
    sheetSelect.innerHTML = '<option value="">Спочатку виберіть Excel-файл</option>';
    previewMeta.textContent = 'Файл ще не вибрано.';
    processStatus.textContent = 'Можна прикріпити тільки .xlsx або .xlsm файл.';
    return;
  }
  try {
    ensureExcelFileSizeAllowed(file);
  } catch (error) {
    fileInput.value = '';
    updateAttachedFileState(null);
    sheetSelect.innerHTML = '<option value="">Спочатку виберіть Excel-файл</option>';
    previewMeta.textContent = 'Файл ще не вибрано.';
    processStatus.textContent = error.message;
    return;
  }
  try {
    workbookState = await analyzeWorkbook(file);
    sheetSelect.innerHTML = '';
    workbookState.sheets.forEach((sheet) => {
      const option = document.createElement('option');
      option.value = sheet.name;
      option.textContent = sheet.name;
      sheetSelect.appendChild(option);
    });
    const savedSheet = loadSettings().sheet_name;
    if (savedSheet && workbookState.sheets.some((sheet) => sheet.name === savedSheet)) {
      sheetSelect.value = savedSheet;
    } else if (workbookState.sheets.length) {
      sheetSelect.value = workbookState.sheets[0].name;
    }
    saveSettings();
    await renderSelectedPreview();
  } catch (error) {
    workbookState = null;
    sheetSelect.innerHTML = '<option value="">Не вдалося прочитати аркуші</option>';
    previewMeta.textContent = error.message;
  }
});

async function loadArticleRequirements() {
  if (!workbookState) {
    throw new Error('Спочатку виберіть Excel-файл.');
  }
  const parsed = await parseSheet(sheetSelect.value);
  const articleColumn = columnToIndex(processForm.elements.article_column.value);
  const startRow = Math.max(1, Number(processForm.elements.start_row.value || 1));
  const requirements = [];
  const seen = new Set();
  for (let row = startRow; row <= parsed.maxRow; row += 1) {
    const value = parsed.rows.get(row)?.get(articleColumn) || '';
    const key = normalizeLocalArticle(value, true);
    if (key && !seen.has(key)) {
      seen.add(key);
      requirements.push({ key });
    }
  }
  return requirements;
}

function selectLocalImages(requirements, order) {
  const indexes = {
    local_1: buildFolderIndex('local_1'),
    local_2: buildFolderIndex('local_2'),
  };
  const selected = [];
  for (const requirement of requirements) {
    for (const source of order) {
      if (source === 'server') continue;
      const file = indexes[source].get(requirement.key);
      if (file) {
        selected.push({ source, key: requirement.key, file });
        break;
      }
    }
  }
  return selected;
}

function downloadFilename(response) {
  const disposition = response.headers.get('Content-Disposition') || '';
  const encoded = disposition.match(/filename\*=utf-8''([^;]+)/i);
  if (encoded) {
    try {
      return decodeURIComponent(encoded[1]);
    } catch (_) {}
  }
  const quoted = disposition.match(/filename="([^"]+)"/i);
  return quoted ? quoted[1] : 'Excel_with Images.xlsx';
}

function createProgressId() {
  if (window.crypto?.randomUUID) {
    return window.crypto.randomUUID();
  }
  return (
    Date.now().toString(36) +
    '-' +
    Math.random().toString(36).slice(2) +
    '-' +
    Math.random().toString(36).slice(2)
  );
}

function stopProgressPolling() {
  if (progressTimer) {
    window.clearInterval(progressTimer);
    progressTimer = null;
  }
}

function stopClientHeartbeat() {
  if (heartbeatTimer) {
    window.clearInterval(heartbeatTimer);
    heartbeatTimer = null;
  }
}

function sendClientHeartbeat(progressId) {
  if (!progressId || cancelRequested) return;
  fetch('/heartbeat/' + encodeURIComponent(progressId), {
    method: 'POST',
    cache: 'no-store',
    keepalive: true,
  }).catch(() => {});
}

function startClientHeartbeat(progressId) {
  stopClientHeartbeat();
  sendClientHeartbeat(progressId);
  heartbeatTimer = window.setInterval(() => {
    sendClientHeartbeat(progressId);
  }, 5000);
}

function setProcessingUi(isProcessing) {
  submitButton.disabled = isProcessing;
  cancelButton.disabled = !isProcessing;
  if (!isProcessing) {
    stopClientHeartbeat();
    activeProgressId = '';
    cancelRequested = false;
  }
}

function setClientProgress(percent, message, details) {
  const safePercent = Math.max(0, Math.min(100, Number(percent) || 0));
  progressPanel.classList.add('is-visible');
  progressFill.style.width = safePercent + '%';
  progressPercent.textContent = Math.round(safePercent) + '%';
  let text = message || 'Обробка...';
  if (details && details.total) {
    text += ' ' + details.inserted + '/' + details.total;
  }
  progressMessage.textContent = text;
}

function startProgressPolling(progressId, initialMessage) {
  stopProgressPolling();
  setClientProgress(1, initialMessage || 'Передаємо файл на сервер...');
  progressTimer = window.setInterval(async () => {
    try {
      const response = await fetch('/progress/' + encodeURIComponent(progressId), {
        cache: 'no-store',
      });
      const payload = await response.json();
      if (payload && payload.ok) {
        setClientProgress(payload.percent, payload.message, payload);
        if (payload.status === 'done' || payload.status === 'error' || payload.status === 'cancelled') {
          stopProgressPolling();
        }
      }
    } catch (_) {
      // Temporary polling errors are ignored; the main /process request owns the final result.
    }
  }, 700);
}

cancelButton.addEventListener('click', async () => {
  if (!activeProgressId || cancelRequested) return;
  if (!window.confirm('Скасувати поточну обробку?')) return;
  cancelRequested = true;
  cancelButton.disabled = true;
  cancelButton.textContent = 'Скасування...';
  processStatus.textContent = 'Скасовуємо обробку...';
  setClientProgress(100, 'Скасування обробки...');
  try {
    await fetch('/cancel/' + encodeURIComponent(activeProgressId), {
      method: 'POST',
      cache: 'no-store',
    });
  } catch (_) {
    processStatus.textContent = 'Не вдалося надіслати запит скасування.';
    cancelButton.disabled = false;
    cancelRequested = false;
  }
});

function cancelActiveProcessingOnUnload() {
  if (!activeProgressId || cancelRequested) return;
  const url = '/cancel/' + encodeURIComponent(activeProgressId);
  cancelRequested = true;
  stopClientHeartbeat();
  if (navigator.sendBeacon) {
    navigator.sendBeacon(url, new Blob([], { type: 'text/plain' }));
    return;
  }
  fetch(url, {
    method: 'POST',
    cache: 'no-store',
    keepalive: true,
  }).catch(() => {});
}

window.addEventListener('beforeunload', cancelActiveProcessingOnUnload);
window.addEventListener('pagehide', cancelActiveProcessingOnUnload);

processForm.addEventListener('submit', async (event) => {
  event.preventDefault();
  setProcessingUi(true);
  cancelButton.disabled = true;
  submitButton.textContent = 'Обробка...';
  processStatus.textContent = 'Готуємо файл на цьому компʼютері...';
  setClientProgress(1, 'Готуємо файл на цьому компʼютері...');
  hideProcessingReport();
  try {
    const excelFile = fileInput.files[0];
    if (!excelFile) {
      throw new Error('Спочатку виберіть Excel-файл.');
    }
    ensureExcelFileSizeAllowed(excelFile);
    saveSettings();
    columnToIndex(processForm.elements.article_column.value);
    columnToIndex(processForm.elements.image_column.value);
    const order = sourceOrder();
    const requirements = await loadArticleRequirements();
    const selectedLocalImages = selectLocalImages(requirements, order);
    const localBytes = selectedLocalImages.reduce(
      (total, item) => total + item.file.size,
      0
    );
    if (localBytes > maxLocalUploadBytes) {
      throw new Error(
        'Загальний розмір знайдених локальних зображень перевищує ліміт ' +
        (window.EXCEL_IMAGE_SERVER_CONFIG?.maxLocalImagesUploadMb || 500) + ' МБ.'
      );
    }
    processStatus.textContent =
      'Знайдено фото з вибраних папок: ' + selectedLocalImages.length +
      '. Готуємо Excel-файл...';
    setClientProgress(2, 'Готуємо Excel-файл до передачі...');
    const processData = new FormData(processForm);
    processData.delete('file');
    processData.append('excel_filename', excelFile.name);
    processData.append('excel_base64', await fileToBase64(excelFile));
    const manifest = [];
    selectedLocalImages.forEach((item) => {
      processData.append('local_images', item.file, item.file.name);
      manifest.push({ source: item.source, key: item.key });
    });
    processData.append('source_order_json', JSON.stringify(order));
    processData.append('local_image_manifest', JSON.stringify(manifest));
    const progressId = createProgressId();
    activeProgressId = progressId;
    cancelRequested = false;
    cancelButton.disabled = false;
    startClientHeartbeat(progressId);
    startProgressPolling(progressId, 'Передаємо файл на сервер...');
    processData.append('progress_id', progressId);
    const response = await fetch('/process?progress_id=' + encodeURIComponent(progressId), {
      method: 'POST',
      body: processData,
    });
    if (!response.ok) {
      let message = 'Не вдалося обробити файл.';
      try {
        const payload = await response.json();
        message = payload.detail || message;
      } catch (_) {}
      throw new Error(message);
    }

    const result = await response.blob();
    const resultUrl = URL.createObjectURL(result);
    const link = document.createElement('a');
    link.href = resultUrl;
    link.download = downloadFilename(response);
    document.body.appendChild(link);
    link.click();
    link.remove();
    window.setTimeout(() => URL.revokeObjectURL(resultUrl), 1000);
    setClientProgress(100, 'Готово. Починається скачування результату...');
    renderProcessingReport(await loadProcessingReport(progressId), order);
    processStatus.textContent =
      'Готово. Скачування почалося автоматично. Вибраний файл залишився прикріпленим локально.';
  } catch (error) {
    setClientProgress(100, cancelRequested ? 'Обробку скасовано.' : 'Обробку зупинено через помилку.');
    processStatus.textContent = error.message;
  } finally {
    stopProgressPolling();
    setProcessingUi(false);
    submitButton.textContent = 'Обробити файл';
    cancelButton.textContent = 'Скасувати обробку';
  }
});
"""


@asynccontextmanager
async def lifespan(_: FastAPI):
    try:
        load_xml_index(force=True)
    except Exception:
        log("XML index was not loaded at startup.")
    log("Excel Image Server started and ready.")
    yield


app = FastAPI(title="Excel Image Server", lifespan=lifespan)


@app.get("/client.js", include_in_schema=False)
def client_js() -> Response:
    return Response(
        CLIENT_JS,
        media_type="application/javascript; charset=utf-8",
        headers={"Cache-Control": "no-store"},
    )


@app.get("/", response_class=HTMLResponse)
def index_page() -> str:
    return render_page()


@app.get("/status")
def status() -> dict[str, Any]:
    load_xml_index()
    public_xml_meta = {
        key: value
        for key, value in INDEX_META.items()
        if key != "source"
    }
    return {
        "ok": True,
        "busy": PROCESS_LOCK.locked() if PROCESS_LOCK is not None else False,
        "xml": public_xml_meta,
        "config": {
            "port": CONFIG["port"],
            "processing_mode": "single_file_queue",
            "hard_max_output_mb": CONFIG["hard_max_output_mb"],
            "max_upload_mb": CONFIG["max_upload_mb"],
        },
    }


@app.get("/progress/{progress_id}")
def progress(progress_id: str, response: Response) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Pragma"] = "no-cache"
    return get_progress_snapshot(progress_id)


@app.post("/cancel/{progress_id}")
def cancel(progress_id: str, response: Response) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Pragma"] = "no-cache"
    cancelled = request_cancel(progress_id)
    return {
        "ok": cancelled,
        "message": (
            "Скасування запитано."
            if cancelled
            else "Активну обробку для цього запиту не знайдено."
        ),
    }


@app.post("/heartbeat/{progress_id}")
def heartbeat(progress_id: str, response: Response) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Pragma"] = "no-cache"
    return {"ok": mark_client_heartbeat(progress_id)}


@app.get("/reload-index", response_class=HTMLResponse)
def reload_index() -> str:
    load_xml_index(force=True)
    public_xml_meta = {
        key: value
        for key, value in INDEX_META.items()
        if key != "source"
    }
    return render_page(
        "XML-індекс перезавантажено.\n"
        + json.dumps(public_xml_meta, ensure_ascii=False, indent=2)
    )


@app.post("/analyze")
async def analyze(response: Response, request: Request) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Pragma"] = "no-cache"
    workbook_buffer: io.BytesIO | None = None
    try:
        form_data = await request.form(
            max_part_size=CONFIG["max_upload_mb"] * 1024 * 1024 * 2
        )
        workbook_buffer, suffix, _, _ = read_excel_form_to_memory(form_data)
        workbook_info = inspect_workbook(workbook_buffer, suffix=suffix)
    except Exception as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    finally:
        if workbook_buffer is not None:
            workbook_buffer.close()
    return workbook_info


@app.post("/article-requirements")
async def article_requirements(
    response: Response,
    request: Request,
) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, private"
    response.headers["Pragma"] = "no-cache"
    workbook_buffer: io.BytesIO | None = None
    try:
        form_data = await request.form(
            max_part_size=CONFIG["max_upload_mb"] * 1024 * 1024 * 2
        )
        workbook_buffer, suffix, _, _ = read_excel_form_to_memory(form_data)
        return inspect_article_requirements(
            workbook_buffer,
            str(form_data.get("sheet_name", "")),
            str(form_data.get("article_column", "A")),
            int(form_data.get("start_row", 2)),
            suffix=suffix,
        )
    except Exception as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    finally:
        if workbook_buffer is not None:
            workbook_buffer.close()


@app.post("/process")
async def process(request: Request) -> StreamingResponse:
    if PROCESS_LOCK is None:
        raise HTTPException(status_code=503, detail="Сервер ще не готовий.")

    progress_id = normalize_progress_id(request.query_params.get("progress_id", ""))
    set_progress(progress_id, 0, "Очікування черги...")
    mark_progress_requires_heartbeat(progress_id)
    form_data: Any | None = None
    resolved_local_images: list[Any] = []
    lock_acquired = False
    job_dir: Path | None = None
    workbook_buffer: io.BytesIO | None = None
    output_buffer: io.BytesIO | None = None
    try:
        while not await asyncio.to_thread(PROCESS_LOCK.acquire, True, 1):
            raise_if_cancelled(progress_id)
        lock_acquired = True
        set_progress(progress_id, 1, "Отримання Excel-файлу...")

        form_data = await request.form(
            max_part_size=CONFIG["max_upload_mb"] * 1024 * 1024 * 2
        )
        workbook_buffer, input_suffix, input_size, original_filename = (
            read_excel_form_to_memory(form_data)
        )

        resolved_local_images = [
            item for item in form_data.getlist("local_images") if hasattr(item, "file")
        ]
        sheet_name = str(form_data.get("sheet_name", ""))
        article_column = str(form_data.get("article_column", "A"))
        image_column = str(form_data.get("image_column", "B"))
        start_row = int(form_data.get("start_row", 2))
        desired_output_mb = float(form_data.get("desired_output_mb", 20))
        cell_background_color = str(form_data.get("cell_background_color", "D9D9D9"))
        skip_cell_background = str(
            form_data.get("skip_cell_background", "")
        ).lower() in {"true", "1", "yes", "on"}
        resolved_source_order = str(
            form_data.get("source_order_json", '["server","local_1","local_2"]')
        )
        resolved_manifest = str(form_data.get("local_image_manifest", "[]"))

        job_dir = make_job_dir("process")
        set_progress(progress_id, 3, "Підготовка Excel-файлу...")
        set_progress(progress_id, 6, "Підготовка фото...")
        source_order = validate_source_order(resolved_source_order)
        local_images_by_source = save_local_image_uploads(
            resolved_local_images,
            resolved_manifest,
            job_dir,
        )
        for local_image in resolved_local_images:
            close_upload(local_image)
        set_progress(progress_id, 8, "Читання аркушів Excel...")
        result_filename = make_download_filename(original_filename, input_suffix)
        sheets = get_sheet_names(workbook_buffer, suffix=input_suffix)
        selected_sheet = sheet_name.strip() or sheets[0]
        output_buffer = io.BytesIO()

        log("Excel processing started.")
        def report_progress(percent: float, message: str, **extra: Any) -> None:
            set_progress(progress_id, percent, message, **extra)

        raise_if_cancelled(progress_id)
        result = await asyncio.to_thread(
            process_excel,
            input_path=workbook_buffer,
            output_path=output_buffer,
            sheet_name=selected_sheet,
            article_column=article_column,
            image_column=image_column,
            start_row=start_row,
            desired_output_mb=desired_output_mb,
            cell_background_color=cell_background_color,
            use_cell_background=not skip_cell_background,
            local_images_by_source=local_images_by_source,
            source_order=source_order,
            progress_callback=report_progress,
            cancel_callback=lambda: raise_if_cancelled(progress_id),
            input_suffix=input_suffix,
            input_size_bytes=input_size,
            temp_dir=job_dir,
        )
        raise_if_cancelled(progress_id)
        set_progress(progress_id, 96, "Очищення тимчасових файлів...")
        output_bytes = output_buffer.getvalue()
        output_buffer.close()
        output_buffer = None
        workbook_buffer.close()
        workbook_buffer = None
        secure_delete_tree(job_dir)
        job_dir = None
        log(
            "Excel processing finished: "
            f"rows={result['rows_seen']}, "
            f"inserted={result['inserted']}, not_found={len(result['not_found'])}, "
            f"failed={len(result['failed'])}, size_mb={result['output_size_mb']}"
        )
        set_progress(
            progress_id,
            100,
            "Готово. Починається скачування результату...",
            status="done",
            inserted=result["inserted"],
            total=result["inserted"] + len(result["failed"]),
            report={
                "inserted": result["inserted"],
                "not_found_count": len(result["not_found"]),
                "not_found": result["not_found"],
                "failed_count": len(result["failed"]),
                "failed": result["failed"],
                "rows_seen": result["rows_seen"],
            },
        )
        media_type = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if input_suffix == ".xlsm"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return StreamingResponse(
            io.BytesIO(output_bytes),
            media_type=media_type,
            headers={
                "Cache-Control": "no-store, private",
                "Pragma": "no-cache",
                "X-Content-Type-Options": "nosniff",
                "Content-Disposition": f"attachment; filename*=utf-8''{quote_header_filename(result_filename)}",
            },
        )
    except HTTPException:
        set_progress(progress_id, 100, "Обробку зупинено через помилку.", status="error")
        raise
    except ProcessingCancelled as error:
        log("Excel processing cancelled by user.")
        set_progress(progress_id, 100, str(error), status="cancelled")
        raise HTTPException(status_code=499, detail=str(error)) from error
    except Exception as error:
        log("Excel processing failed.")
        set_progress(progress_id, 100, "Обробку зупинено через помилку.", status="error")
        raise HTTPException(status_code=500, detail=str(error)) from error
    finally:
        for local_image in resolved_local_images:
            close_upload(local_image)
        if workbook_buffer is not None:
            workbook_buffer.close()
        if output_buffer is not None:
            output_buffer.close()
        if form_data is not None and hasattr(form_data, "close"):
            close_result = form_data.close()
            if hasattr(close_result, "__await__"):
                await close_result
        if job_dir is not None:
            secure_delete_tree(job_dir)
        if lock_acquired:
            PROCESS_LOCK.release()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Excel image insertion web server.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_FILE)
    parser.add_argument(
        "--print-config-log",
        action="store_true",
        help="Print resolved public log path and exit.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        config = load_config(args.config.resolve())
    except Exception as error:
        print(f"Configuration error: {error}", file=sys.stderr)
        return 1

    if args.print_config_log:
        print(str(config["public_log_file"]))
        return 0

    configure_runtime(config)

    uvicorn.run(
        app,
        host=CONFIG["host"],
        port=CONFIG["port"],
        log_level="info",
        access_log=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
